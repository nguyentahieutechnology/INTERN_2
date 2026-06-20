from datetime import date, datetime, timedelta
from functools import wraps

import openpyxl

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.contrib.auth.views import LoginView
from django.db.models import Count, Max, ProtectedError, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.clickjacking import xframe_options_sameorigin

from core.forms import (DangKyForm, DatLichForm, KhachHangForm,
                        LoVacXinForm, NguoiThanForm, PhacDoChiTietFormSet, PhacDoForm,
                        SuaNhanVienForm, TaiKhamForm, TaiKhoanNhanVienForm,
                        TheoDoiForm, VacXinForm)
from core.models import (KhachHang, LichHen, LichSuQuanTri, LoVacXin,
                         MuiTiem, PhacDo, PhacDoChiTiet, PhieuHuyKho, PhieuSangLoc,
                         PhieuXuatKho, QuyTrinhTiem, ThanhToan, TheoDoiSauTiem,
                         ThongBao, TinNhanHoTro, VacXin, YeuCauDatLaiMatKhau)

# Nguong tuoi (thang) phan loai vac-xin tre em / nguoi lon
TUOI_TRE_EM_MAX_THANG = 108   # < 9 tuoi
TUOI_NGUOI_LON_MIN_THANG = 216  # >= 18 tuoi


def dang_ky(request):
    """Trang khach hang dang ky tai khoan moi (cong khai)."""
    if request.user.is_authenticated:
        return redirect('sau_dang_nhap')
    if request.method == 'POST':
        form = DangKyForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)            # tu dong dang nhap sau khi dang ky
            return redirect('ca_nhan')
    else:
        form = DangKyForm()
    return render(request, 'core/dang_ky.html', {'form': form})


def quen_mat_khau(request):
    """Khach nhap ten dang nhap -> gui yeu cau dat lai mat khau cho admin."""
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        user = User.objects.filter(username=username).first()
        if not user:
            messages.error(request, f'Không tìm thấy tài khoản "{username}".')
        else:
            yc, created = YeuCauDatLaiMatKhau.objects.get_or_create(
                user=user, da_xu_ly=False)
            if created:
                for admin in User.objects.filter(is_superuser=True, is_active=True):
                    ThongBao.objects.create(
                        nguoi_nhan=admin,
                        noi_dung=f'Yêu cầu đặt lại mật khẩu từ tài khoản "{username}".',
                        duong_dan=reverse('quan_tri_yeu_cau_mat_khau'))
            messages.success(request,
                             'Đã gửi yêu cầu tới quản trị. Vui lòng chờ được cấp lại '
                             'mật khẩu rồi đăng nhập lại.')
            return redirect('dang_nhap')
    return render(request, 'core/quen_mat_khau.html')


@login_required
def sau_dang_nhap(request):
    """Dieu huong sau dang nhap theo vai tro."""
    u = request.user
    nhom = set(u.groups.values_list('name', flat=True))
    if 'Le tan' in nhom:
        return redirect('le_tan_dashboard')
    if 'Bac si' in nhom:
        return redirect('bac_si_dashboard')
    if 'Dieu duong' in nhom:
        return redirect('dieu_duong_dashboard')
    if 'Thu kho' in nhom:
        return redirect('thu_kho_dashboard')
    if u.is_staff or u.is_superuser:
        return redirect('dashboard')          # admin -> bang dieu khien
    return redirect('index')                  # khach hang -> trang chu


@login_required
def ca_nhan(request):
    """Danh sach ho so (ban than + nguoi than) cua tai khoan."""
    ho_so_list = request.user.ho_so.all().order_by('id')
    if not ho_so_list:
        return redirect('dashboard')
    return render(request, 'core/ca_nhan.html', {'ho_so_list': ho_so_list})


@login_required
def ho_so_chi_tiet(request, pk):
    """Chi tiet 1 ho so: so tiem + lich hen + theo doi (phai thuoc tai khoan)."""
    kh = get_object_or_404(KhachHang, pk=pk, user=request.user)
    mui_list = kh.mui_tiem.select_related('vac_xin', 'lo', 'nguoi_tiem').order_by('ngay_tiem', 'mui_so')
    mui_moi_nhat = kh.mui_tiem.order_by('-ngay_tiem', '-mui_so').first()
    context = {
        'kh': kh,
        'mui_list': mui_list,
        'lich_hen_list': kh.lich_hen.all(),
        'theo_doi_list': TheoDoiSauTiem.objects.filter(mui_tiem__khach_hang=kh),
        'co_mui_tiem': mui_list.exists(),
        'ngay_nhac': mui_moi_nhat.ngay_hen_mui_ke if mui_moi_nhat else None,
        'today': timezone.now().date(),
    }
    return render(request, 'core/ho_so_chi_tiet.html', context)


@login_required
@xframe_options_sameorigin
def phieu_tiem_pdf(request, pk):
    """Phieu xac nhan tiem chung cho 1 mui tiem (xuat tu ho so khach)."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle

    mt = get_object_or_404(MuiTiem.objects.select_related(
        'khach_hang', 'vac_xin', 'lo', 'nguoi_tiem'), pk=pk)
    kh = mt.khach_hang
    if kh.user != request.user and not request.user.is_staff:
        return render(request, 'core/khong_co_quyen.html', status=403)
    font = _font_pdf_dam()

    def ten(u):
        return (u.get_full_name() or u.username) if u else '—'

    s_ten = ParagraphStyle('sten', fontName=font, fontSize=11)
    s_td = ParagraphStyle('std', fontName=font, fontSize=15, alignment=1, spaceBefore=4, spaceAfter=8)
    s_dong = ParagraphStyle('sdong', fontName=font, fontSize=10.5, spaceAfter=4)
    s_lbl = ParagraphStyle('slbl', fontName=font, fontSize=10)
    s_val = ParagraphStyle('sval', fontName=font, fontSize=10)
    s_muc = ParagraphStyle('smuc', fontName=font, fontSize=10.5, spaceBefore=8, spaceAfter=4,
                           textColor=colors.HexColor('#0d6efd'))
    s_ky = ParagraphStyle('sky', fontName=font, fontSize=10, alignment=1, spaceBefore=4)
    s_ky_it = ParagraphStyle('skyit', fontName=font, fontSize=8.5, alignment=1, textColor=colors.grey)
    s_tenky = ParagraphStyle('stenky', fontName=font, fontSize=10, alignment=1, spaceBefore=18)

    lo_txt = (f'{mt.lo.so_lo} (HSD {mt.lo.han_su_dung:%d/%m/%Y})') if mt.lo else '—'
    mui_ke = mt.ngay_hen_mui_ke.strftime('%d/%m/%Y') if mt.ngay_hen_mui_ke else '—'

    el = [
        Paragraph('PHÒNG KHÁM TIÊM CHỦNG AN TÂM', s_ten),
        Paragraph('PHIẾU XÁC NHẬN TIÊM CHỦNG', s_td),
        Paragraph(f'Họ và tên: <b>{kh.ho_ten}</b>', s_dong),
        Paragraph(f'Ngày sinh: {kh.ngay_sinh:%d/%m/%Y}    -    Giới tính: {kh.gioi_tinh}', s_dong),
    ]

    rows = [
        ['Vắc-xin', f'{mt.vac_xin.ten} (phòng {mt.vac_xin.phong_benh})'],
        ['Mũi số', str(mt.mui_so)],
        ['Ngày tiêm', mt.ngay_tiem.strftime('%d/%m/%Y')],
        ['Số lô (HSD)', lo_txt],
        ['Vị trí tiêm', mt.vi_tri_tiem or '—'],
        ['Người tiêm', ten(mt.nguoi_tiem)],
        ['Hẹn mũi kế tiếp', mui_ke],
    ]
    data = [[Paragraph(a, s_lbl), Paragraph(b, s_val)] for a, b in rows]
    t = Table(data, colWidths=[4.5 * cm, 8.5 * cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#adb5bd')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    el.append(t)

    theo_doi = list(mt.theo_doi.all())
    el.append(Paragraph('Theo dõi sau tiêm', s_muc))
    if theo_doi:
        for td_ in theo_doi:
            el.append(Paragraph(
                f'- {td_.get_thoi_diem_display()}: {td_.get_muc_do_display()}'
                + (f' (nhiệt độ {td_.nhiet_do}°C)' if td_.nhiet_do else '')
                + (f' — {td_.trieu_chung}' if td_.trieu_chung else ''), s_dong))
    else:
        el.append(Paragraph('- Chưa ghi nhận theo dõi.', s_dong))

    el.append(Spacer(1, 0.5 * cm))
    khoi = [Paragraph('Người tiêm', s_ky), Paragraph('(ký, ghi rõ họ tên)', s_ky_it),
            Paragraph(ten(mt.nguoi_tiem), s_tenky)]
    t_ky = Table([['', khoi]], colWidths=[6.5 * cm, 6.5 * cm])
    t_ky.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    el.append(t_ky)

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=A5, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                      leftMargin=1.4 * cm, rightMargin=1.4 * cm).build(el)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    cach = 'attachment' if request.GET.get('tai') else 'inline'
    resp['Content-Disposition'] = f'{cach}; filename=phieu_tiem_{mt.id}.pdf'
    return resp


@login_required
@xframe_options_sameorigin
def phieu_tiem_xem(request, pk):
    """Phieu tiem dang trang HTML (giong to phieu) - hien trong modal, khong tai ve."""
    mt = get_object_or_404(MuiTiem.objects.select_related(
        'khach_hang', 'vac_xin', 'lo', 'nguoi_tiem'), pk=pk)
    kh = mt.khach_hang
    if kh.user != request.user and not request.user.is_staff:
        return render(request, 'core/khong_co_quyen.html', status=403)
    # Bac si: lay tu buoc sang loc cua quy trinh gan voi mui tiem nay
    bac_si = None
    qt = mt.quytrinhtiem_set.select_related('lich_hen__sang_loc__bac_si').first()
    if qt and hasattr(qt.lich_hen, 'sang_loc'):
        bac_si = qt.lich_hen.sang_loc.bac_si
    return render(request, 'core/phieu_tiem.html', {
        'mt': mt, 'kh': kh, 'theo_doi': mt.theo_doi.all(), 'bac_si': bac_si,
    })


@login_required
def khach_huy_lich(request, lich_id):
    """Khach tu huy lich hen cua minh (khi con cho xac nhan / da xac nhan)."""
    lh = get_object_or_404(LichHen, pk=lich_id, khach_hang__user=request.user)
    if request.method == 'POST' and lh.trang_thai in ('cho', 'xacnhan'):
        lh.trang_thai = 'huy'
        lh.save()
        _bao_le_tan(f'Khách "{lh.khach_hang.ho_ten}" đã tự hủy lịch hẹn '
                    f'ngày {lh.ngay_hen:%d/%m/%Y}.', reverse('le_tan_lich_hen'))
        messages.success(request, 'Đã hủy lịch hẹn.')
    return redirect('ho_so_chi_tiet', pk=lh.khach_hang_id)


@login_required
def khach_sua_ho_so(request, pk):
    """Khach tu cap nhat thong tin ho so cua minh (ban than / nguoi than)."""
    kh = get_object_or_404(KhachHang, pk=pk, user=request.user)
    if request.method == 'POST':
        form = KhachHangForm(request.POST, instance=kh)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật thông tin.')
            return redirect('ho_so_chi_tiet', pk=kh.id)
    else:
        form = KhachHangForm(instance=kh)
    return render(request, 'core/ho_so_sua.html', {'form': form, 'kh': kh})


@login_required
@xframe_options_sameorigin
def so_tiem_pdf(request, pk):
    """Xuat so tiem chung / giay chung nhan tiem chung (PDF)."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle

    kh = get_object_or_404(KhachHang, pk=pk)
    if kh.user != request.user and not request.user.is_staff:
        return render(request, 'core/khong_co_quyen.html', status=403)

    font = _font_pdf()
    today = timezone.now().date()
    mui = kh.mui_tiem.select_related('vac_xin', 'lo', 'nguoi_tiem').order_by('ngay_tiem', 'mui_so')

    tieu_de = ParagraphStyle('td', fontName=font, fontSize=15, alignment=1, spaceAfter=4)
    phu = ParagraphStyle('phu', fontName=font, fontSize=9, alignment=1,
                         textColor=colors.grey, spaceAfter=2)
    thuong = ParagraphStyle('th', fontName=font, fontSize=10, spaceAfter=2)

    el = [
        Paragraph('SỔ TIÊM CHỦNG ĐIỆN TỬ', tieu_de),
        Paragraph('Phòng khám Tiêm chủng An Tâm', phu),
        Paragraph(f'Ngày cấp: {today:%d/%m/%Y}', phu),
        Spacer(1, 0.4 * cm),
        Paragraph(f'Họ và tên: {kh.ho_ten}', thuong),
        Paragraph(f'Ngày sinh: {kh.ngay_sinh:%d/%m/%Y}    |    '
                  f'Giới tính: {kh.gioi_tinh}    |    Điện thoại: {kh.so_dien_thoai}', thuong),
        Spacer(1, 0.3 * cm),
    ]

    data = [['Mũi', 'Vắc-xin', 'Phòng bệnh', 'Ngày tiêm', 'Số lô', 'Người tiêm']]
    for m in mui:
        nguoi = (m.nguoi_tiem.get_full_name() or m.nguoi_tiem.username) if m.nguoi_tiem else '-'
        data.append([str(m.mui_so), m.vac_xin.ten, m.vac_xin.phong_benh,
                     m.ngay_tiem.strftime('%d/%m/%Y'),
                     m.lo.so_lo if m.lo else '-', nguoi])
    if len(data) == 1:
        data.append(['', 'Chưa có mũi tiêm nào', '', '', '', ''])

    t = Table(data, repeatRows=1,
              colWidths=[1.2 * cm, 4.5 * cm, 4 * cm, 2.6 * cm, 2.2 * cm, 3.3 * cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (3, 0), (4, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#adb5bd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    el.append(t)
    el.append(Spacer(1, 0.6 * cm))
    el.append(Paragraph('Tổng số mũi đã tiêm: ' + str(mui.count()), thuong))

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                      leftMargin=1.5 * cm, rightMargin=1.5 * cm).build(el)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    cach = 'attachment' if request.GET.get('tai') else 'inline'
    resp['Content-Disposition'] = f'{cach}; filename=so_tiem_{kh.id}.pdf'
    return resp


def _doc_so_tien(so):
    """Doc so tien (VND) thanh chu tieng Viet."""
    so = int(so or 0)
    if so == 0:
        return 'Không đồng'
    dv = ['', 'một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín']

    def doc3(n, day_du):
        tram, chuc, donvi = n // 100, (n % 100) // 10, n % 10
        s = ''
        if tram > 0 or day_du:
            s += (dv[tram] if tram > 0 else 'không') + ' trăm'
        if chuc > 1:
            s += ' ' + dv[chuc] + ' mươi'
            if donvi == 1:
                s += ' mốt'
            elif donvi == 5:
                s += ' lăm'
            elif donvi > 0:
                s += ' ' + dv[donvi]
        elif chuc == 1:
            s += ' mười'
            if donvi == 5:
                s += ' lăm'
            elif donvi > 0:
                s += ' ' + dv[donvi]
        elif donvi > 0:
            s += (' lẻ ' if s else '') + dv[donvi]
        return s.strip()

    nhom = []
    while so > 0:
        nhom.append(so % 1000)
        so //= 1000
    nhom.reverse()
    don_vi_nhom = ['', ' nghìn', ' triệu', ' tỷ']
    n = len(nhom)
    kq = ''
    for i, g in enumerate(nhom):
        if g == 0:
            continue
        kq += ' ' + doc3(g, i != 0) + don_vi_nhom[n - 1 - i]
    kq = kq.strip()
    return kq[0].upper() + kq[1:] + ' đồng'


@login_required
@xframe_options_sameorigin
def phieu_thanh_toan_pdf(request, pk):
    """Xuat phieu thu tien (PDF) theo mau co header + so tien bang chu + 2 cot ky."""
    import os
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A5, landscape
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle

    tt = get_object_or_404(ThanhToan.objects.select_related('khach_hang', 'lich_hen__vac_xin'), pk=pk)
    kh = tt.khach_hang
    if (kh.user != request.user) and not request.user.is_staff:
        return render(request, 'core/khong_co_quyen.html', status=403)
    font = _font_pdf()
    # Dang ky font dam (de dung <b>) neu chua co
    if font == 'VN' and 'VN-Bold' not in pdfmetrics.getRegisteredFontNames():
        for bp in (r'C:\Windows\Fonts\arialbd.ttf', r'C:\Windows\Fonts\segoeuib.ttf',
                   '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'):
            if os.path.exists(bp):
                try:
                    pdfmetrics.registerFont(TTFont('VN-Bold', bp))
                    pdfmetrics.registerFontFamily('VN', normal='VN', bold='VN-Bold',
                                                  italic='VN', boldItalic='VN-Bold')
                except Exception:
                    pass
                break

    def vnd(x):
        return format(int(x or 0), ',d')

    s_ten = ParagraphStyle('sten', fontName=font, fontSize=11, alignment=0)
    s_phai = ParagraphStyle('sphai', fontName=font, fontSize=9)
    s_td = ParagraphStyle('std', fontName=font, fontSize=15, alignment=1, spaceBefore=4, spaceAfter=10)
    s_dong = ParagraphStyle('sdong', fontName=font, fontSize=10.5, spaceAfter=5)
    s_ngay = ParagraphStyle('sngay', fontName=font, fontSize=10, alignment=1, spaceBefore=8)
    s_ky = ParagraphStyle('sky', fontName=font, fontSize=10.5, alignment=1)
    s_ky_it = ParagraphStyle('skyit', fontName=font, fontSize=8.5, alignment=1, textColor=colors.grey)
    s_tenky = ParagraphStyle('stenky', fontName=font, fontSize=10, alignment=1, spaceBefore=18)
    s_luuy = ParagraphStyle('sluuy', fontName=font, fontSize=8.5, spaceAfter=2)
    s_luuyh = ParagraphStyle('sluuyh', fontName=font, fontSize=9, spaceBefore=12, spaceAfter=2)

    nguoi = tt.nguoi_thu or (request.user if request.user.is_staff else None)
    ten_thu = (nguoi.get_full_name() or nguoi.username) if nguoi else ''
    noi_dung = tt.lich_hen.vac_xin.ten if tt.lich_hen and tt.lich_hen.vac_xin else 'tiêm vắc-xin'
    now = timezone.localtime(tt.ngay_thanh_toan)

    def hang2(trai, phai):
        t2 = Table([[Paragraph(trai, s_dong), Paragraph(phai, s_dong)]],
                   colWidths=[11.5 * cm, 6.5 * cm])
        t2.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'),
                                ('LEFTPADDING', (0, 0), (-1, -1), 0)]))
        return t2

    el = []
    hdr = Table([[Paragraph('PHÒNG KHÁM TIÊM CHỦNG AN TÂM', s_ten),
                  Paragraph(f'Quyển số: TC001<br/>Số CT: {tt.id}', s_phai)]],
                colWidths=[13 * cm, 5 * cm])
    hdr.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    el.append(hdr)
    el.append(Paragraph('PHIẾU THU TIỀN', s_td))

    el.append(hang2(f'- Họ tên bệnh nhân : <b>{kh.ho_ten}</b>', f'<b>Năm Sinh: {kh.ngay_sinh:%Y}</b>'))
    el.append(hang2(f'- Địa chỉ : {kh.dia_chi or "................"}', 'Khoa: Phòng tiêm vắc-xin'))
    el.append(Paragraph(f'- Thu tiền : thu tiền {noi_dung}', s_dong))
    el.append(Paragraph(f'- Số tiền : <b>{vnd(tt.tong_tien)}</b>', s_dong))
    el.append(Paragraph(f'- <b>Viết bằng chữ : {_doc_so_tien(tt.tong_tien)}.</b>', s_dong))
    t_ngay = Table([['', Paragraph(f'Ngày {now:%d} tháng {now:%m} năm {now:%Y}', s_ngay)]],
                   colWidths=[9 * cm, 9 * cm])
    t_ngay.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    el.append(t_ngay)

    cot_nop = [Paragraph('NGƯỜI NỘP TIỀN', s_ky), Paragraph('(Ký, ghi rõ họ tên)', s_ky_it),
               Paragraph(kh.ho_ten, s_tenky)]
    cot_thu = [Paragraph('NGƯỜI THU TIỀN', s_ky), Paragraph('(Ký, ghi rõ họ tên)', s_ky_it),
               Paragraph(ten_thu, s_tenky)]
    t_ky = Table([[cot_nop, cot_thu]], colWidths=[9 * cm, 9 * cm])
    t_ky.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    el.append(t_ky)

    el.append(Paragraph('Lưu ý :', s_luuyh))
    for d in ('Phiếu chỉ lưu hành trong phòng khám. Thời gian thanh toán 3 ngày từ lúc kết thúc khám.',
              'Bệnh nhân lưu giữ phiếu thu này trong suốt thời gian khám và điều trị.',
              'Phòng khám không chịu trách nhiệm nếu bệnh nhân làm mất phiếu.'):
        el.append(Paragraph('- ' + d, s_luuy))

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=landscape(A5), topMargin=1 * cm, bottomMargin=1 * cm,
                      leftMargin=1.3 * cm, rightMargin=1.3 * cm).build(el)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    cach = 'attachment' if request.GET.get('tai') else 'inline'
    resp['Content-Disposition'] = f'{cach}; filename=phieu_thu_{tt.id}.pdf'
    return resp


@login_required
def them_nguoi_than(request):
    """Them ho so nguoi than vao tai khoan."""
    if request.method == 'POST':
        form = NguoiThanForm(request.POST)
        if form.is_valid():
            kh = form.save(commit=False)
            kh.user = request.user
            kh.save()
            return redirect('ca_nhan')
    else:
        form = NguoiThanForm()
    return render(request, 'core/them_nguoi_than.html', {'form': form})


@login_required
def xem_thong_bao(request, pk):
    """Danh dau thong bao da doc roi chuyen den trang lien quan."""
    tb = get_object_or_404(ThongBao, pk=pk, nguoi_nhan=request.user)
    if not tb.da_doc:
        tb.da_doc = True
        tb.save(update_fields=['da_doc'])
    # Co duong dan -> chuyen den; khong co -> chi danh dau da doc va o lai trang hien tai
    return redirect(tb.duong_dan or request.META.get('HTTP_REFERER') or 'index')


@login_required
def doc_het_thong_bao(request):
    """Danh dau tat ca thong bao da doc."""
    ThongBao.objects.filter(nguoi_nhan=request.user, da_doc=False).update(da_doc=True)
    return redirect(request.META.get('HTTP_REFERER', 'index'))


# ===================== HO TRO TRUC TUYEN (CHAT) =====================

@login_required
def ho_tro_tin(request):
    """JSON: hoi thoai ho tro cua khach hang hien tai (danh dau tin NV da doc)."""
    qs = TinNhanHoTro.objects.filter(khach=request.user)
    qs.filter(la_nhan_vien=True, da_doc=False).update(da_doc=True)
    tin = [{'noi_dung': t.noi_dung, 'nv': t.la_nhan_vien,
            'gio': timezone.localtime(t.ngay).strftime('%H:%M %d/%m')}
           for t in qs]
    return JsonResponse({'tin': tin})


@login_required
def ho_tro_gui(request):
    """Khach gui tin nhan ho tro."""
    if request.method == 'POST':
        nd = (request.POST.get('noi_dung') or '').strip()
        if nd:
            TinNhanHoTro.objects.create(
                khach=request.user, nguoi_gui=request.user, noi_dung=nd[:1000])
            return JsonResponse({'ok': True})
    return JsonResponse({'ok': False}, status=400)


@login_required
def theo_doi_sau_tiem(request):
    """Khach hang ghi nhan theo doi sau tiem (cho mui tiem cua ho so minh quan ly)."""
    mui_qs = MuiTiem.objects.filter(khach_hang__user=request.user)
    if request.method == 'POST':
        form = TheoDoiForm(request.POST, mui_queryset=mui_qs)
        if form.is_valid():
            form.save()
            return redirect('ca_nhan')
    else:
        form = TheoDoiForm(mui_queryset=mui_qs)
    return render(request, 'core/theo_doi.html', {
        'form': form, 'co_mui_tiem': mui_qs.exists()})


@login_required
def dat_lich(request):
    """Khach hang dat lich hen tiem cho ho so minh quan ly."""
    ho_so = request.user.ho_so.all()
    if not ho_so:
        return redirect('dashboard')
    if request.method == 'POST':
        form = DatLichForm(request.POST, ho_so=ho_so)
        if form.is_valid():
            lich = form.save(commit=False)
            lich.trang_thai = 'cho'           # cho xac nhan
            lich.save()
            return redirect('ca_nhan')
    else:
        form = DatLichForm(ho_so=ho_so)
    # Map phac_do -> danh sach vac_xin id (de loc goi theo vac-xin khi dat lich)
    pd_map = {
        str(pd.id): sorted(set(pd.chi_tiet.values_list('vac_xin_id', flat=True)))
        for pd in PhacDo.objects.prefetch_related('chi_tiet')
    }
    return render(request, 'core/dat_lich.html', {
        'form': form, 'phac_do_vacxin': pd_map})

# So giay giu phien khi "ghi nho dang nhap" (14 ngay)
GHI_NHO_SECONDS = 14 * 24 * 60 * 60


class DangNhapView(LoginView):
    """Trang dang nhap rieng, ho tro 'Ghi nho dang nhap'."""
    template_name = 'core/dang_nhap.html'

    def form_valid(self, form):
        if self.request.POST.get('ghi_nho'):
            # Giu phien 14 ngay du dong trinh duyet
            self.request.session.set_expiry(GHI_NHO_SECONDS)
        else:
            # Het phien khi dong trinh duyet
            self.request.session.set_expiry(0)
        return super().form_valid(form)


def index(request):
    """Trang index cong khai - khach vao thay dau tien (khong can dang nhap)."""
    context = {
        'danh_sach_vacxin': VacXin.objects.all().order_by('ten'),
        'so_vacxin': VacXin.objects.count(),
    }
    return render(request, 'core/index.html', context)


# ===== Cac trang cong khai cua thanh menu phu =====

def dich_vu(request):
    """Trang gioi thieu dich vu."""
    return render(request, 'core/dich_vu.html')


def vac_xin_tre_em(request):
    """Danh muc vac-xin cho tre em."""
    ds = VacXin.objects.filter(
        do_tuoi_min_thang__lt=TUOI_TRE_EM_MAX_THANG).order_by('ten')
    return render(request, 'core/vac_xin_list.html', {
        'ds': ds,
        'tieu_de': 'Vắc xin cho trẻ em',
        'mo_ta': 'Các loại vắc-xin phù hợp cho trẻ sơ sinh và trẻ nhỏ.',
    })


def vac_xin_nguoi_lon(request):
    """Danh muc vac-xin cho nguoi lon."""
    ds = VacXin.objects.filter(
        do_tuoi_max_thang__gte=TUOI_NGUOI_LON_MIN_THANG).order_by('ten')
    return render(request, 'core/vac_xin_list.html', {
        'ds': ds,
        'tieu_de': 'Vắc xin cho người lớn',
        'mo_ta': 'Các loại vắc-xin dành cho thanh thiếu niên và người trưởng thành.',
    })


def goi_vac_xin(request):
    """Tat ca goi tiem / phac do."""
    ds = PhacDo.objects.prefetch_related('chi_tiet').all()
    return render(request, 'core/goi_vac_xin.html', {
        'ds': ds, 'tieu_de': 'Gói vắc xin',
        'mo_ta': 'Các gói tiêm trọn liệu trình theo độ tuổi và nhu cầu.'})


def goi_vac_xin_tre_em(request):
    """Goi tiem cho tre em."""
    ds = PhacDo.objects.filter(nhom='tre_em').prefetch_related('chi_tiet')
    return render(request, 'core/goi_vac_xin.html', {
        'ds': ds, 'tieu_de': 'Gói vắc xin cho trẻ em',
        'mo_ta': 'Các gói tiêm dành cho trẻ sơ sinh và trẻ nhỏ.'})


def goi_vac_xin_nguoi_lon(request):
    """Goi tiem cho nguoi lon."""
    ds = PhacDo.objects.filter(nhom='nguoi_lon').prefetch_related('chi_tiet')
    return render(request, 'core/goi_vac_xin.html', {
        'ds': ds, 'tieu_de': 'Gói vắc xin cho người lớn',
        'mo_ta': 'Các gói tiêm dành cho thanh thiếu niên và người trưởng thành.'})


def cam_nang(request):
    """Cam nang tiem chung (noi dung tinh)."""
    return render(request, 'core/cam_nang.html')


def bang_gia(request):
    """Bang gia vac-xin."""
    ds = VacXin.objects.all().order_by('ten')
    return render(request, 'core/bang_gia.html', {'ds': ds})


def trong_nhom(*ten_nhom):
    """Decorator: chi cho phep superuser hoac user thuoc 1 trong cac nhom."""
    def decorator(view):
        @wraps(view)
        @login_required
        def wrapper(request, *args, **kwargs):
            u = request.user
            if u.is_superuser or u.groups.filter(name__in=ten_nhom).exists():
                return view(request, *args, **kwargs)
            return render(request, 'core/khong_co_quyen.html',
                          {'nhom_can': ten_nhom}, status=403)
        return wrapper
    return decorator


@login_required
def dashboard(request):
    """Dashboard tong hop cho admin: bieu do + so lieu Le tan, Thu kho, Nhac mui."""
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('ca_nhan')            # khach hang -> ho so ca nhan
    today = timezone.now().date()

    # ---------- Le tan: doanh thu + bieu do 6 thang ----------
    tong_dt = ThanhToan.objects.aggregate(s=Sum('tong_tien'))['s'] or 0
    dt_thang = (ThanhToan.objects
                .filter(ngay_thanh_toan__year=today.year,
                        ngay_thanh_toan__month=today.month)
                .aggregate(s=Sum('tong_tien'))['s'] or 0)
    nhan, dt_list, tiem_list = [], [], []
    for i in range(5, -1, -1):
        mm, yy = today.month - i, today.year
        while mm <= 0:
            mm += 12
            yy -= 1
        nhan.append(f'{mm:02d}/{yy}')
        dt_list.append(int(
            ThanhToan.objects.filter(ngay_thanh_toan__year=yy, ngay_thanh_toan__month=mm)
            .aggregate(s=Sum('tong_tien'))['s'] or 0))
        tiem_list.append(
            MuiTiem.objects.filter(ngay_tiem__year=yy, ngay_tiem__month=mm).count())

    # ---------- Thu kho ----------
    los = LoVacXin.objects.select_related('vac_xin')
    han_canh_bao = today + timedelta(days=60)

    # ---------- Nhac mui ----------
    nhac_full = _danh_sach_nhac_mui(None, today)            # tat ca khach can nhac
    so_qua_han = sum(1 for r in nhac_full if r['so_ngay'] < 0)
    so_sap_toi = sum(1 for r in nhac_full if 0 <= r['so_ngay'] < 30)
    nhac_top = nhac_full[:8]                                # da sort theo ngay_hen

    context = {
        'today': today,
        # Tong quan
        'so_khach': KhachHang.objects.count(),
        'so_mui_tiem': MuiTiem.objects.count(),
        'so_den_han': sum(1 for r in nhac_full if 0 <= r['so_ngay'] <= 7),
        # Le tan
        'tong_doanh_thu': tong_dt,
        'doanh_thu_thang': dt_thang,
        'cho_xac_nhan': LichHen.objects.filter(trang_thai='cho').count(),
        'tong_luot_tiem': MuiTiem.objects.count(),
        'chart_nhan': nhan,
        'chart_doanh_thu': dt_list,
        'chart_luot_tiem': tiem_list,
        # Thu kho
        'tong_ton': los.aggregate(s=Sum('so_luong_ton'))['s'] or 0,
        'cho_xuat_kho': (QuyTrinhTiem.objects
                         .filter(giai_doan=QuyTrinhTiem.B_XUAT_KHO)
                         .exclude(lich_hen__trang_thai='huy').count()),
        'so_sap_het_han': los.filter(so_luong_ton__gt=0, han_su_dung__gte=today,
                                     han_su_dung__lte=han_canh_bao).count(),
        'so_het_han': los.filter(so_luong_ton__gt=0, han_su_dung__lt=today).count(),
        'so_ton_thap': los.filter(so_luong_ton__gt=0, so_luong_ton__lt=20).count(),
        # Nhac mui
        'so_qua_han': so_qua_han,
        'so_sap_toi': so_sap_toi,
        'nhac_top': nhac_top,
    }
    return render(request, 'core/trang_chu.html', context)


@trong_nhom('Quan tri', 'Le tan', 'Dieu duong')
def bao_cao_nhac_mui(request):
    """
    Bao cao khach den han nhac mui. Chi Quan tri / Le tan / Dieu duong xem duoc
    (Bac si tap trung vao sang loc -> bi tu choi 403).
    Voi moi khach, lay mui tiem MOI NHAT; neu mui do co ngay_hen_mui_ke
    va <= nguong (mac dinh 7 ngay toi) thi dua vao danh sach nhac.
    """
    today = timezone.now().date()
    days_raw = request.GET.get('days', '7')
    if days_raw == 'all':
        so_ngay = 'all'
        han_chot = None                       # khong gioi han thoi gian
    else:
        try:
            so_ngay = int(days_raw)
        except (TypeError, ValueError):
            so_ngay = 7
        han_chot = today + timedelta(days=so_ngay)

    danh_sach = _danh_sach_nhac_mui(han_chot, today)
    qua_han = [r for r in danh_sach if r['so_ngay'] < 0]
    sap_toi = [r for r in danh_sach if 0 <= r['so_ngay'] < 30]    # duoi 1 thang

    context = {
        'danh_sach': danh_sach,
        'so_ngay': so_ngay,
        'today': today,
        'tong': len(danh_sach),
        'so_qua_han': len(qua_han),
        'so_sap_toi': len(sap_toi),
    }
    return render(request, 'core/nhac_mui.html', context)


def _danh_sach_nhac_mui(han_chot, today):
    """Tra ve list dict cac khach can nhac mui, sap xep theo ngay hen."""
    ket_qua = []
    for kh in KhachHang.objects.all():
        # mui tiem moi nhat cua khach
        mui_moi_nhat = kh.mui_tiem.order_by('-ngay_tiem', '-mui_so').first()
        if not mui_moi_nhat or not mui_moi_nhat.ngay_hen_mui_ke:
            continue
        if han_chot is not None and mui_moi_nhat.ngay_hen_mui_ke > han_chot:
            continue
        so_ngay = (mui_moi_nhat.ngay_hen_mui_ke - today).days
        ket_qua.append({
            'khach_hang': kh,
            'vac_xin': mui_moi_nhat.vac_xin,
            'mui_ke': mui_moi_nhat.mui_so + 1,
            'ngay_hen': mui_moi_nhat.ngay_hen_mui_ke,
            'so_ngay': so_ngay,
            'qua_han': so_ngay < 0,
        })
    ket_qua.sort(key=lambda r: r['ngay_hen'])
    return ket_qua


def chung_nhan_tiem(request, pk):
    """Trang cong khai hien thi chung nhan tiem chung (cho phep quet QR xem nhanh)."""
    kh = get_object_or_404(KhachHang, pk=pk)
    mui_list = kh.mui_tiem.select_related('vac_xin', 'lo').order_by('ngay_tiem', 'mui_so')
    return render(request, 'core/ho_so_cong_khai.html', {
        'kh': kh,
        'mui_list': mui_list,
        'today': timezone.now()
    })


# ===================== KHU VUC LE TAN =====================

@trong_nhom('Quan tri', 'Le tan')
def le_tan_dashboard(request):
    """Dashboard tong hop: doanh thu, khach hang, luot tiem + bieu do 6 thang."""
    today = timezone.now().date()
    tong_dt = ThanhToan.objects.aggregate(s=Sum('tong_tien'))['s'] or 0
    dt_thang = (ThanhToan.objects
                .filter(ngay_thanh_toan__year=today.year,
                        ngay_thanh_toan__month=today.month)
                .aggregate(s=Sum('tong_tien'))['s'] or 0)
    khach_tiem = (KhachHang.objects.annotate(so_mui=Count('mui_tiem'))
                  .filter(so_mui__gt=0).order_by('-so_mui')[:20])

    # Du lieu bieu do 6 thang gan nhat
    nhan, dt_thang_list, tiem_thang_list = [], [], []
    for i in range(5, -1, -1):
        mm, yy = today.month - i, today.year
        while mm <= 0:
            mm += 12
            yy -= 1
        nhan.append(f'{mm:02d}/{yy}')
        dt_thang_list.append(int(
            ThanhToan.objects.filter(ngay_thanh_toan__year=yy, ngay_thanh_toan__month=mm)
            .aggregate(s=Sum('tong_tien'))['s'] or 0))
        tiem_thang_list.append(
            MuiTiem.objects.filter(ngay_tiem__year=yy, ngay_tiem__month=mm).count())

    top_kh = list(khach_tiem)[:8]
    
    # Canh bao kho: Sap het han (1 thang) hoac Ton thap (< 10)
    han_canh_bao = today + timedelta(days=30)
    lo_sap_het_han = LoVacXin.objects.filter(so_luong_ton__gt=0, han_su_dung__lte=han_canh_bao).order_by('han_su_dung')
    lo_ton_thap = LoVacXin.objects.filter(so_luong_ton__gt=0, so_luong_ton__lt=10).order_by('so_luong_ton')

    context = {
        'chart_nhan': nhan,
        'chart_doanh_thu': dt_thang_list,
        'chart_luot_tiem': tiem_thang_list,
        'chart_kh_nhan': [k.ho_ten for k in top_kh],
        'chart_kh_somui': [k.so_mui for k in top_kh],
        'tong_doanh_thu': tong_dt,
        'doanh_thu_thang': dt_thang,
        'tong_khach': KhachHang.objects.count(),
        'tong_luot_tiem': MuiTiem.objects.count(),
        'cho_xac_nhan': LichHen.objects.filter(trang_thai='cho').count(),
        'khach_tiem': khach_tiem,
        'thanh_toan_moi': (ThanhToan.objects
                           .select_related('khach_hang').order_by('-ngay_thanh_toan')),
        'today': today,
        'nam_list': list(range(today.year - 9, today.year + 1)),
        'nam_hien': today.year,
        'thang_hien': today.strftime('%Y-%m'),
        'lo_sap_het_han': lo_sap_het_han,
        'lo_ton_thap': lo_ton_thap,
    }
    return render(request, 'core/le_tan/dashboard.html', context)


def _du_lieu_thong_ke(che_do, nam_param, thang_param, today):
    """Doanh thu + luot tiem theo nam/thang/ngay.
    Tra: labels, dt_list, lt_list, mo_ta (tieu de), nhan_cot (ten cot dau bang)."""
    import calendar

    def dt_f(**kw):
        return int(ThanhToan.objects.filter(**kw).aggregate(s=Sum('tong_tien'))['s'] or 0)

    labels, dt_list, lt_list = [], [], []
    if che_do == 'nam':
        for y in range(today.year - 9, today.year + 1):
            labels.append(str(y))
            dt_list.append(dt_f(ngay_thanh_toan__year=y))
            lt_list.append(MuiTiem.objects.filter(ngay_tiem__year=y).count())
        return labels, dt_list, lt_list, '10 năm gần nhất', 'Năm'
    if che_do == 'ngay':
        try:
            yy, mm = map(int, (thang_param or today.strftime('%Y-%m')).split('-'))
        except (ValueError, AttributeError):
            yy, mm = today.year, today.month
        for d in range(1, calendar.monthrange(yy, mm)[1] + 1):
            labels.append(f'{d:02d}')
            dt_list.append(dt_f(ngay_thanh_toan__year=yy, ngay_thanh_toan__month=mm,
                                ngay_thanh_toan__day=d))
            lt_list.append(MuiTiem.objects.filter(
                ngay_tiem__year=yy, ngay_tiem__month=mm, ngay_tiem__day=d).count())
        return labels, dt_list, lt_list, f'Các ngày trong tháng {mm:02d}/{yy}', 'Ngày'
    try:
        yy = int(nam_param or today.year)
    except (TypeError, ValueError):
        yy = today.year
    for m in range(1, 13):
        labels.append(f'{m:02d}/{yy}')
        dt_list.append(dt_f(ngay_thanh_toan__year=yy, ngay_thanh_toan__month=m))
        lt_list.append(MuiTiem.objects.filter(ngay_tiem__year=yy, ngay_tiem__month=m).count())
    return labels, dt_list, lt_list, f'12 tháng năm {yy}', 'Tháng'


@trong_nhom('Quan tri', 'Le tan')
def le_tan_thong_ke(request):
    """API JSON: doanh thu + luot tiem + lich su thanh toan theo nam / thang / ngay."""
    from django.http import JsonResponse
    che_do = request.GET.get('che_do', 'thang')
    nam_p, thang_p = request.GET.get('nam'), request.GET.get('thang')
    today = timezone.now().date()
    labels, dt_list, lt_list, _, _ = _du_lieu_thong_ke(che_do, nam_p, thang_p, today)

    # Lich su thanh toan theo dung khoang thoi gian dang chon
    tt_qs = ThanhToan.objects.select_related('khach_hang').order_by('-ngay_thanh_toan')
    if che_do == 'nam':
        tt_qs = tt_qs.filter(ngay_thanh_toan__year__gte=today.year - 9)
    elif che_do == 'ngay':
        try:
            yy, mm = map(int, (thang_p or today.strftime('%Y-%m')).split('-'))
        except (ValueError, AttributeError):
            yy, mm = today.year, today.month
        tt_qs = tt_qs.filter(ngay_thanh_toan__year=yy, ngay_thanh_toan__month=mm)
    else:
        try:
            yy = int(nam_p or today.year)
        except (TypeError, ValueError):
            yy = today.year
        tt_qs = tt_qs.filter(ngay_thanh_toan__year=yy)

    def vnd(x):
        return format(int(x or 0), ',d').replace(',', '.')
    thanh_toan = [{
        'ngay': timezone.localtime(tt.ngay_thanh_toan).strftime('%d/%m/%Y %H:%M'),
        'khach': tt.khach_hang.ho_ten,
        'pt': tt.get_phuong_thuc_display(),
        'tien': vnd(tt.tong_tien),
    } for tt in tt_qs]

    return JsonResponse({'labels': labels, 'doanh_thu': dt_list, 'luot_tiem': lt_list,
                         'thanh_toan': thanh_toan})


@trong_nhom('Quan tri', 'Le tan')
@xframe_options_sameorigin
def le_tan_bao_cao_pdf(request):
    """Xuat bao cao hoat dong le tan ra PDF (doanh thu, khach, luot tiem)."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import (VerticalBarChart,
                                                     HorizontalBarChart)
    from reportlab.graphics.charts.linecharts import HorizontalLineChart

    today = timezone.now().date()
    font = _font_pdf()

    def vnd(x):
        return format(int(x or 0), ',d').replace(',', '.')

    tong_dt = ThanhToan.objects.aggregate(s=Sum('tong_tien'))['s'] or 0
    dt_thang = (ThanhToan.objects
                .filter(ngay_thanh_toan__year=today.year,
                        ngay_thanh_toan__month=today.month)
                .aggregate(s=Sum('tong_tien'))['s'] or 0)
    tong_khach = KhachHang.objects.count()
    tong_luot = MuiTiem.objects.count()
    cho_xn = LichHen.objects.filter(trang_thai='cho').count()

    # Du lieu bieu do theo dung bo loc dang chon tren dashboard
    ky_labels, dt_vals, lt_vals, mo_ta, nhan_cot = _du_lieu_thong_ke(
        request.GET.get('che_do', 'thang'), request.GET.get('nam'),
        request.GET.get('thang'), today)
    dt_trieu = [round(v / 1e6, 2) for v in dt_vals]

    top_kh = (KhachHang.objects.annotate(so_mui=Count('mui_tiem'))
              .filter(so_mui__gt=0).order_by('-so_mui')[:10])

    tieu_de = ParagraphStyle('td', fontName=font, fontSize=15, alignment=1, spaceAfter=4)
    phu = ParagraphStyle('phu', fontName=font, fontSize=9, alignment=1,
                         textColor=colors.grey, spaceAfter=2)
    thuong = ParagraphStyle('th', fontName=font, fontSize=10, spaceAfter=2)
    muc = ParagraphStyle('muc', fontName=font, fontSize=11, spaceBefore=10,
                         spaceAfter=4, textColor=colors.HexColor('#0d6efd'))

    def bang(data, widths):
        t = Table(data, repeatRows=1, colWidths=widths)
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#adb5bd')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        return t

    def _truc_font(ch):
        ch.categoryAxis.labels.fontName = font
        ch.categoryAxis.labels.fontSize = 7
        ch.valueAxis.labels.fontName = font
        ch.valueAxis.labels.fontSize = 7

    def cot(nhan, vals, mau):
        dr = Drawing(470, 165)
        bc = VerticalBarChart()
        bc.x, bc.y, bc.width, bc.height = 45, 25, 410, 125
        bc.data = [vals]
        bc.categoryAxis.categoryNames = nhan
        bc.valueAxis.valueMin = 0
        bc.bars[0].fillColor = mau
        _truc_font(bc)
        dr.add(bc)
        return dr

    def duong(nhan, vals, mau):
        dr = Drawing(470, 165)
        lc = HorizontalLineChart()
        lc.x, lc.y, lc.width, lc.height = 45, 25, 410, 125
        lc.data = [vals]
        lc.categoryAxis.categoryNames = nhan
        lc.valueAxis.valueMin = 0
        lc.lines[0].strokeColor = mau
        lc.lines[0].strokeWidth = 2
        _truc_font(lc)
        dr.add(lc)
        return dr

    def thanh_ngang(ten, vals, mau):
        n = max(1, len(ten))
        dr = Drawing(470, 24 * n + 30)
        bc = HorizontalBarChart()
        bc.x, bc.y, bc.width, bc.height = 130, 15, 320, 24 * n
        bc.data = [vals]
        bc.categoryAxis.categoryNames = ten
        bc.valueAxis.valueMin = 0
        bc.bars[0].fillColor = mau
        _truc_font(bc)
        dr.add(bc)
        return dr

    nguoi = request.user.get_full_name() or request.user.username
    el = [
        Paragraph('BÁO CÁO HOẠT ĐỘNG – LỄ TÂN', tieu_de),
        Paragraph('Phòng khám Tiêm chủng An Tâm', phu),
        Paragraph(f'Ngày xuất: {today:%d/%m/%Y}', phu),
        Paragraph(f'Người xuất báo cáo: {nguoi}', phu),
        Spacer(1, 0.4 * cm),
        Paragraph(f'Tổng doanh thu: {vnd(tong_dt)} đ', thuong),
        Paragraph(f'Doanh thu tháng {today.month}/{today.year}: {vnd(dt_thang)} đ', thuong),
        Paragraph(f'Tổng khách hàng: {tong_khach}    |    Tổng lượt tiêm: {tong_luot}'
                  f'    |    Lịch chờ xác nhận: {cho_xn}', thuong),
        Paragraph(f'Biểu đồ doanh thu — {mo_ta} (triệu đồng)', muc),
        cot(ky_labels, dt_trieu, colors.HexColor('#198754')),
        Paragraph(f'Biểu đồ lượt tiêm — {mo_ta}', muc),
        duong(ky_labels, lt_vals, colors.HexColor('#0d6efd')),
        Paragraph(f'Chi tiết theo {nhan_cot}', muc),
    ]
    d1 = [[nhan_cot, 'Doanh thu (đ)', 'Lượt tiêm']]
    for ten, dtv, ltv in zip(ky_labels, dt_vals, lt_vals):
        d1.append([ten, vnd(dtv), str(ltv)])
    el.append(bang(d1, [4 * cm, 7 * cm, 4 * cm]))

    el.append(Paragraph('Khách hàng tiêm nhiều nhất', muc))
    if top_kh:
        ten_kh = [k.ho_ten for k in top_kh][::-1]
        somui = [k.so_mui for k in top_kh][::-1]
        el.append(thanh_ngang(ten_kh, somui, colors.HexColor('#0dcaf0')))
    d2 = [['STT', 'Khách hàng', 'Điện thoại', 'Số lần tiêm']]
    for i, kh in enumerate(top_kh, 1):
        d2.append([str(i), kh.ho_ten, kh.so_dien_thoai, str(kh.so_mui)])
    if not top_kh:
        d2.append(['', 'Chưa có lượt tiêm', '', ''])
    el.append(bang(d2, [1.4 * cm, 7 * cm, 3.8 * cm, 2.8 * cm]))

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                      leftMargin=1.5 * cm, rightMargin=1.5 * cm).build(el)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    cach = 'attachment' if request.GET.get('tai') else 'inline'
    resp['Content-Disposition'] = f'{cach}; filename=bao_cao_le_tan_{today:%Y%m%d}.pdf'
    return resp


@trong_nhom('Quan tri', 'Le tan', 'Dieu duong')
def nhac_khach(request, pk):
    """Gui thong bao chuong nhac mui cho khach hang."""
    kh = get_object_or_404(KhachHang, pk=pk)
    if request.method == 'POST':
        if kh.user:
            mui = kh.mui_tiem.order_by('-ngay_tiem', '-mui_so').first()
            ngay = mui.ngay_hen_mui_ke if mui else None
            noi_dung = f'Nhắc lịch tiêm cho {kh.ho_ten}: mũi tiếp theo'
            if ngay:
                noi_dung += f' dự kiến {ngay:%d/%m/%Y}'
            noi_dung += '. Vui lòng đặt lịch hẹn.'
            ThongBao.objects.create(nguoi_nhan=kh.user, noi_dung=noi_dung,
                                    duong_dan=reverse('dat_lich'))
            messages.success(request, f'Đã gửi nhắc cho {kh.ho_ten}.')
        else:
            messages.warning(request, f'{kh.ho_ten} chưa có tài khoản để nhận thông báo.')
    return redirect(request.META.get('HTTP_REFERER') or 'bao_cao_nhac_mui')


@trong_nhom('Quan tri', 'Le tan')
def le_tan_lich_hen(request):
    """Danh sach + loc lich hen."""
    trang_thai = request.GET.get('trang_thai', '')
    q = request.GET.get('q', '').strip()
    ds = LichHen.objects.select_related('khach_hang', 'phac_do').order_by('-ngay_hen', '-id')
    if trang_thai:
        ds = ds.filter(trang_thai=trang_thai)
    if q:
        ds = ds.filter(khach_hang__ho_ten__icontains=q)
    return render(request, 'core/le_tan/lich_hen.html', {
        'ds': ds, 'trang_thai': trang_thai, 'q': q,
        'cac_trang_thai': LichHen.TRANG_THAI,
    })


@trong_nhom('Quan tri', 'Le tan')
def le_tan_dang_ky_tiem(request):
    """Le tan dang ky tiem (dat lich) cho khach tai quay -> vao hang cho.
    Tai quay: ngay/gio hen mac dinh la hom nay/bay gio."""
    khach = KhachHang.objects.all().order_by('ho_ten')
    if request.method == 'POST':
        form = DatLichForm(request.POST, ho_so=khach)
        if form.is_valid():
            lich = form.save(commit=False)
            lich.trang_thai = 'cho'           # vao hang cho de xac nhan
            lich.save()
            return redirect('le_tan_lich_hen')
    else:
        bay_gio = timezone.localtime()
        form = DatLichForm(ho_so=khach, initial={
            'ngay_hen': bay_gio.date(),
            'gio_hen': bay_gio.time().replace(second=0, microsecond=0),
        })
    pd_map = {
        str(pd.id): sorted(set(pd.chi_tiet.values_list('vac_xin_id', flat=True)))
        for pd in PhacDo.objects.prefetch_related('chi_tiet')
    }
    # Ton kho con han su dung theo tung vac-xin (de bao "het hang" ngay khi chon)
    today = timezone.now().date()
    ton_qs = (LoVacXin.objects.filter(han_su_dung__gte=today)
              .values('vac_xin_id').annotate(s=Sum('so_luong_ton')))
    vacxin_ton = {str(r['vac_xin_id']): r['s'] or 0 for r in ton_qs}
    return render(request, 'core/le_tan/dang_ky_tiem.html', {
        'form': form, 'phac_do_vacxin': pd_map, 'vacxin_ton': vacxin_ton})


@trong_nhom('Quan tri', 'Le tan')
def le_tan_tai_kham(request):
    """Dat lich tai kham sau phan ung (form rieng, don gian)."""
    khach = KhachHang.objects.all().order_by('ho_ten')
    if request.method == 'POST':
        form = TaiKhamForm(request.POST, ho_so=khach)
        if form.is_valid():
            lich = form.save(commit=False)
            lich.trang_thai = 'cho'
            # Danh dau tai kham sau phan ung (de to do + nhan dien trong lich hen)
            if 'Tái khám sau phản ứng' not in (lich.ghi_chu or ''):
                lich.ghi_chu = ('Tái khám sau phản ứng tiêm. ' + (lich.ghi_chu or '')).strip()
            lich.save()
            messages.success(request, 'Đã đặt lịch tái khám.')
            return redirect('le_tan_lich_hen')
    else:
        bay_gio = timezone.localtime()
        initial = {
            'ngay_hen': bay_gio.date(),
            'gio_hen': bay_gio.time().replace(second=0, microsecond=0),
        }
        if request.GET.get('khach_hang'):
            initial['khach_hang'] = request.GET['khach_hang']
        if request.GET.get('ghi_chu'):
            initial['ghi_chu'] = request.GET['ghi_chu']
        form = TaiKhamForm(ho_so=khach, initial=initial)
    return render(request, 'core/le_tan/tai_kham.html', {'form': form})


@trong_nhom('Quan tri', 'Le tan')
def le_tan_doi_trang_thai(request, pk, trang_thai):
    """Doi trang thai lich hen (signal se tu bao cho khach)."""
    if request.method == 'POST' and trang_thai in dict(LichHen.TRANG_THAI):
        lh = get_object_or_404(LichHen, pk=pk)
        lh.trang_thai = trang_thai
        lh.save()
        if trang_thai == 'xacnhan':
            # Bat dau quy trinh tiem khi xac nhan lich
            QuyTrinhTiem.objects.get_or_create(lich_hen=lh)
    return redirect(request.META.get('HTTP_REFERER') or 'le_tan_lich_hen')


@trong_nhom('Quan tri', 'Le tan')
def le_tan_da_den(request, pk):
    """Le tan xac nhan khach da den -> sang quy trinh, nhay buoc 2 (dang cho).
    Dung chung cho ca lich tiem thuong va tai kham."""
    lh = get_object_or_404(LichHen, pk=pk)
    if request.method == 'POST':
        if lh.trang_thai == 'cho':
            lh.trang_thai = 'xacnhan'
            lh.save()
        qt, _ = QuyTrinhTiem.objects.get_or_create(lich_hen=lh)
        if qt.giai_doan < QuyTrinhTiem.B_DANG_CHO:
            qt.giai_doan = QuyTrinhTiem.B_DANG_CHO
            qt.save()
        return redirect('quy_trinh_chi_tiet', pk=qt.id)
    return redirect('le_tan_lich_hen')


@trong_nhom('Quan tri', 'Le tan')
def le_tan_ho_tro(request):
    """Hop thoai ho tro: danh sach khach + tra loi."""
    if request.method == 'POST':
        khach = User.objects.filter(pk=request.POST.get('khach')).first()
        nd = (request.POST.get('noi_dung') or '').strip()
        if khach and nd:
            TinNhanHoTro.objects.create(
                khach=khach, nguoi_gui=request.user, la_nhan_vien=True, noi_dung=nd[:1000])
            ThongBao.objects.create(
                nguoi_nhan=khach,
                noi_dung='Phòng khám đã trả lời tin nhắn hỗ trợ của bạn.',
                duong_dan=reverse('ca_nhan'))
        return redirect(f"{reverse('le_tan_ho_tro')}?khach={khach.pk if khach else ''}")

    threads = (User.objects.filter(tin_ho_tro__isnull=False).distinct()
               .annotate(chua_doc=Count('tin_ho_tro',
                         filter=Q(tin_ho_tro__la_nhan_vien=False, tin_ho_tro__da_doc=False)),
                         tin_cuoi=Max('tin_ho_tro__ngay'))
               .order_by('-tin_cuoi'))
    chon = User.objects.filter(pk=request.GET.get('khach')).first()
    tin = []
    if chon:
        TinNhanHoTro.objects.filter(khach=chon, la_nhan_vien=False, da_doc=False).update(da_doc=True)
        tin = TinNhanHoTro.objects.filter(khach=chon)
    return render(request, 'core/le_tan/ho_tro.html',
                  {'threads': threads, 'chon': chon, 'tin': tin})


@trong_nhom('Quan tri', 'Le tan')
def le_tan_khach_hang(request):
    """Danh sach + tim kiem + them khach hang (tiep don tai quay)."""
    if request.method == 'POST':
        form = KhachHangForm(request.POST)
        if form.is_valid():
            form.save()                       # user=None, quan_he mac dinh
            return redirect('le_tan_khach_hang')
    else:
        form = KhachHangForm()
    q = request.GET.get('q', '').strip()
    ds = KhachHang.objects.all().order_by('-ngay_tao')
    if q:
        ds = ds.filter(Q(ho_ten__icontains=q) | Q(so_dien_thoai__icontains=q))
    return render(request, 'core/le_tan/khach_hang.html', {'ds': ds, 'q': q, 'form': form})


@trong_nhom('Quan tri', 'Le tan')
def le_tan_khach_hang_sua(request, pk):
    """Sua thong tin khach hang (mo khi bam dup vao dong)."""
    kh = get_object_or_404(KhachHang, pk=pk)
    if request.method == 'POST':
        form = KhachHangForm(request.POST, instance=kh)
        if form.is_valid():
            form.save()
            return redirect('le_tan_khach_hang')
    else:
        form = KhachHangForm(instance=kh)
    return render(request, 'core/le_tan/khach_hang_sua.html', {'form': form, 'kh': kh})


# ===================== QUY TRINH TIEM =====================

def _timeline(qt):
    """Cac buoc hien thi tien trinh. Tai kham: 3 buoc tiep don + bac si kham."""
    if qt.la_tai_kham:
        buoc = [(QuyTrinhTiem.B_DA_TOI, 'Khách đã tới'),
                (QuyTrinhTiem.B_DANG_CHO, 'Khách đang chờ'),
                (QuyTrinhTiem.B_VAO_PHONG, 'Đã vào phòng khám'),
                (QuyTrinhTiem.B_SANG_LOC, 'Bác sĩ khám')]
    else:
        buoc = [(ma, ten) for ma, ten in QuyTrinhTiem.GIAI_DOAN
                if ma != QuyTrinhTiem.HOAN_TAT]
    return [{'so': ma, 'ten': ten,
             'xong': qt.giai_doan > ma,
             'hien_tai': qt.giai_doan == ma}
            for ma, ten in buoc]


@trong_nhom('Quan tri', 'Le tan', 'Bac si', 'Dieu duong', 'Thu kho')
def quy_trinh_list(request):
    """Danh sach quy trinh tiem (cac lich da xac nhan tro di)."""
    ds = (QuyTrinhTiem.objects
          .select_related('lich_hen__khach_hang', 'vac_xin')
          .exclude(lich_hen__trang_thai='huy')      # bo lich da huy
          .order_by('giai_doan', '-ngay_cap_nhat'))
    return render(request, 'core/le_tan/quy_trinh_list.html', {'ds': ds})


@trong_nhom('Quan tri', 'Le tan', 'Bac si', 'Dieu duong', 'Thu kho')
def quy_trinh_chi_tiet(request, pk):
    """Chi tiet 1 quy trinh: tien trinh + thao tac theo buoc."""
    qt = get_object_or_404(QuyTrinhTiem.objects.select_related(
        'lich_hen__khach_hang', 'vac_xin', 'lo', 'mui_tiem', 'thanh_toan'), pk=pk)
    kh = qt.lich_hen.khach_hang
    theo_doi = qt.mui_tiem.theo_doi.all() if qt.mui_tiem else []
    context = {
        'qt': qt,
        'kh': kh,
        'sang_loc': getattr(qt.lich_hen, 'sang_loc', None),
        'phieu_xuat': qt.phieu_xuat.first(),
        'ly_do_chong_choices': PhieuSangLoc.LY_DO_CHONG,
        'timeline': _timeline(qt),
        'los': (LoVacXin.objects.filter(vac_xin=qt.vac_xin, so_luong_ton__gt=0)
                if qt.vac_xin else LoVacXin.objects.none()),
        'theo_doi': theo_doi,
        'today': timezone.now().date(),
    }
    return render(request, 'core/le_tan/quy_trinh_chi_tiet.html', context)


@trong_nhom('Quan tri', 'Le tan')
def qt_tiep_don(request, pk):
    """Le tan: tiep don 3 buoc (da toi -> dang cho -> vao phong)."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if request.method == 'POST' and qt.giai_doan in (
            QuyTrinhTiem.B_DA_TOI, QuyTrinhTiem.B_DANG_CHO, QuyTrinhTiem.B_VAO_PHONG):
        qt.giai_doan += 1
        qt.save()
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Bac si')
def qt_sang_loc(request, pk):
    """Bac si kham sang loc -> ghi PhieuSangLoc. Dat -> len don, hoan/chong -> huy lich."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if request.method == 'POST' and qt.giai_doan == QuyTrinhTiem.B_SANG_LOC:
        ket_luan = request.POST.get('ket_luan', 'dat')
        ghi_chu = request.POST.get('ghi_chu', '')
        dung_thuoc = bool(request.POST.get('dang_dung_thuoc'))
        PhieuSangLoc.objects.update_or_create(
            lich_hen=qt.lich_hen,
            defaults=dict(
                bac_si=request.user,
                nhiet_do=request.POST.get('nhiet_do') or None,
                mach=request.POST.get('mach') or None,
                huyet_ap=request.POST.get('huyet_ap', ''),
                tien_su_benh=request.POST.get('tien_su_benh', ''),
                di_ung=request.POST.get('di_ung', ''),
                benh_di_truyen=request.POST.get('benh_di_truyen', ''),
                dang_dung_thuoc=dung_thuoc,
                thuoc_dang_dung=request.POST.get('thuoc_dang_dung', '') if dung_thuoc else '',
                ket_luan=ket_luan, ghi_chu=ghi_chu,
                ly_do_hoan=request.POST.get('ly_do_hoan', '') if ket_luan == 'hoan' else '',
                ly_do_chong=request.POST.get('ly_do_chong', '') if ket_luan == 'chong' else '',
                ly_do_chong_khac=(request.POST.get('ly_do_chong_khac', '')
                                  if ket_luan == 'chong' and request.POST.get('ly_do_chong') == 'khac' else '')))
        qt.ghi_chu_sang_loc = ghi_chu
        if ket_luan == 'dat':
            # Du dieu kien -> lay vac-xin khach da dat tu lich hen, sang thanh toan
            qt.vac_xin = qt.lich_hen.vac_xin
            qt.giai_doan = QuyTrinhTiem.B_THANH_TOAN
            qt.save()
        else:
            # Hoan / chong chi dinh -> huy lich, dung quy trinh
            qt.save()
            qt.lich_hen.trang_thai = 'huy'
            qt.lich_hen.save()
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Bac si')
def qt_kham(request, pk):
    """Bac si kham ca tai kham sau phan ung -> ghi ket qua va hoan tat."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if (request.method == 'POST' and qt.la_tai_kham
            and qt.giai_doan == QuyTrinhTiem.B_SANG_LOC):
        qt.ghi_chu_sang_loc = request.POST.get('ket_qua', '')
        qt.giai_doan = QuyTrinhTiem.HOAN_TAT
        qt.save()
        qt.lich_hen.trang_thai = 'datiem'      # 'datiem' -> badge hien "Da kham" cho tai kham
        qt.lich_hen.save()
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Le tan')
def qt_thanh_toan(request, pk):
    """Le tan thu tien -> tao ban ghi ThanhToan."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if request.method == 'POST' and qt.giai_doan == QuyTrinhTiem.B_THANH_TOAN and qt.vac_xin:
        tt = ThanhToan.objects.create(
            khach_hang=qt.lich_hen.khach_hang, lich_hen=qt.lich_hen,
            tong_tien=qt.vac_xin.gia,
            phuong_thuc=request.POST.get('phuong_thuc', 'tien_mat'),
            nguoi_thu=request.user)
        qt.thanh_toan = tt
        qt.giai_doan = QuyTrinhTiem.B_XUAT_KHO
        qt.save()
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Thu kho')
def qt_xuat_kho(request, pk):
    """Thu kho xuat kho (chon lo, tru ton)."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    lo_id = request.POST.get('lo')
    if request.method == 'POST' and qt.giai_doan == QuyTrinhTiem.B_XUAT_KHO and lo_id:
        lo = get_object_or_404(LoVacXin, pk=lo_id)
        if lo.so_luong_ton > 0:
            lo.so_luong_ton -= 1
            lo.save()
            qt.lo = lo
            qt.giai_doan = QuyTrinhTiem.B_TIEM
            qt.save()
            PhieuXuatKho.objects.create(lo=lo, so_luong=1,
                                        nguoi_xuat=request.user, quy_trinh=qt)
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Dieu duong')
def qt_nhan_vacxin(request, pk):
    """Dieu duong xac nhan da nhan vac-xin tu kho (hoan tat chu ky phieu xuat)."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if request.method == 'POST' and qt.giai_doan == QuyTrinhTiem.B_TIEM:
        px = qt.phieu_xuat.first()
        if px and not px.nguoi_nhan:
            px.nguoi_nhan = request.user
            px.ngay_nhan = timezone.now()
            px.save(update_fields=['nguoi_nhan', 'ngay_nhan'])
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Dieu duong')
def qt_tiem(request, pk):
    """Dieu duong ghi nhan tiem -> tao MuiTiem."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if request.method == 'POST' and qt.giai_doan == QuyTrinhTiem.B_TIEM:
        kh = qt.lich_hen.khach_hang
        mui_so = MuiTiem.objects.filter(khach_hang=kh, vac_xin=qt.vac_xin).count() + 1
        mt = MuiTiem.objects.create(
            khach_hang=kh, vac_xin=qt.vac_xin, lo=qt.lo,
            phac_do=qt.lich_hen.phac_do, mui_so=mui_so,
            ngay_tiem=timezone.now().date(), nguoi_tiem=request.user,
            vi_tri_tiem=request.POST.get('vi_tri_tiem', ''))
        qt.mui_tiem = mt
        qt.giai_doan = QuyTrinhTiem.B_THEO_DOI
        qt.save()
        qt.lich_hen.trang_thai = 'datiem'
        qt.lich_hen.save()
    return redirect('quy_trinh_chi_tiet', pk=pk)


@trong_nhom('Quan tri', 'Dieu duong')
def qt_theo_doi(request, pk):
    """Ghi nhan theo doi sau tiem -> hoan tat quy trinh."""
    qt = get_object_or_404(QuyTrinhTiem, pk=pk)
    if request.method == 'POST' and qt.giai_doan == QuyTrinhTiem.B_THEO_DOI and qt.mui_tiem:
        TheoDoiSauTiem.objects.create(
            mui_tiem=qt.mui_tiem,
            thoi_diem=request.POST.get('thoi_diem', '30p'),
            nhiet_do=request.POST.get('nhiet_do') or None,
            trieu_chung=request.POST.get('trieu_chung', ''),
            muc_do=request.POST.get('muc_do', 'binh_thuong'))
        qt.giai_doan = QuyTrinhTiem.HOAN_TAT
        qt.save()
    return redirect('quy_trinh_chi_tiet', pk=pk)


# ============= TRANG LAM VIEC RIENG THEO VAI TRO =============

@trong_nhom('Quan tri', 'Bac si')
def bac_si_dashboard(request):
    """Hang cho cua bac si: cho sang loc / duyet cho tiem."""
    cho_sang_loc = (QuyTrinhTiem.objects
                    .select_related('lich_hen__khach_hang', 'lich_hen__vac_xin')
                    .filter(giai_doan=QuyTrinhTiem.B_SANG_LOC)
                    .exclude(lich_hen__trang_thai='huy'))
    return render(request, 'core/bac_si/dashboard.html', {
        'cho_sang_loc': cho_sang_loc,
    })


@trong_nhom('Quan tri', 'Bac si')
def bac_si_lich_su(request):
    """Lich su kham sang loc."""
    q = request.GET.get('q', '').strip()
    ds = (PhieuSangLoc.objects
          .select_related('lich_hen__khach_hang', 'bac_si')
          .order_by('-ngay'))
    if q:
        ds = ds.filter(lich_hen__khach_hang__ho_ten__icontains=q)
    return render(request, 'core/bac_si/lich_su.html', {'ds': ds, 'q': q})


@trong_nhom('Quan tri', 'Dieu duong')
def dieu_duong_dashboard(request):
    """Hang cho cua dieu duong: cho tiem + cho theo doi + canh bao phan ung."""
    base = (QuyTrinhTiem.objects
            .select_related('lich_hen__khach_hang', 'vac_xin', 'mui_tiem')
            .exclude(lich_hen__trang_thai='huy'))
    canh_bao = (TheoDoiSauTiem.objects
                .filter(muc_do__in=['can_chu_y', 'nghiem_trong'])
                .select_related('mui_tiem__khach_hang', 'mui_tiem__vac_xin')[:10])
    return render(request, 'core/dieu_duong/dashboard.html', {
        'cho_tiem': base.filter(giai_doan=QuyTrinhTiem.B_TIEM),
        'cho_theo_doi': base.filter(giai_doan=QuyTrinhTiem.B_THEO_DOI),
        'canh_bao': canh_bao,
        'xu_ly_choices': TheoDoiSauTiem.XU_LY,
    })


@trong_nhom('Quan tri', 'Dieu duong')
def dieu_duong_xu_ly_phan_ung(request, pk):
    """Dieu duong gui yeu cau xu ly cho ca canh bao phan ung."""
    from urllib.parse import urlencode
    td = get_object_or_404(
        TheoDoiSauTiem.objects.select_related('mui_tiem__khach_hang'), pk=pk)
    if request.method == 'POST':
        xu_ly = request.POST.get('xu_ly', '')
        ly_do = request.POST.get('xu_ly_ghi_chu', '').strip()
        if xu_ly in dict(TheoDoiSauTiem.XU_LY):
            td.xu_ly = xu_ly
            td.xu_ly_ghi_chu = ly_do
            td.save(update_fields=['xu_ly', 'xu_ly_ghi_chu'])
            kh = td.mui_tiem.khach_hang
            if xu_ly == 'quay_lai':
                # Bao le tan -> bam vao chuyen sang dat lich (dien san khach + ly do)
                ghi_chu = 'Tái khám sau phản ứng tiêm'
                if ly_do:
                    ghi_chu += f': {ly_do[:120]}'
                duong_dan = (reverse('le_tan_tai_kham') + '?'
                             + urlencode({'khach_hang': kh.id, 'ghi_chu': ghi_chu}))
                _bao_le_tan(f'Khách "{kh.ho_ten}" cần quay lại tái khám sau phản ứng tiêm.'
                            + (f' Lý do: {ly_do}.' if ly_do else '')
                            + ' Bấm để đặt lịch.', duong_dan)
                messages.success(request, f'Đã chuyển yêu cầu tái khám của "{kh.ho_ten}" cho lễ tân.')
            else:
                # Phan hoi truc tiep cho khach
                noi_dung = ly_do if (xu_ly == 'khac' and ly_do) else td.get_xu_ly_display()
                if kh.user:
                    ThongBao.objects.create(
                        nguoi_nhan=kh.user,
                        noi_dung=f'Phản hồi theo dõi sau tiêm: {noi_dung}.',
                        duong_dan='')
                messages.success(request, f'Đã ghi nhận xử lý cho "{kh.ho_ten}".')
    return redirect('dieu_duong_dashboard')


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_dashboard(request):
    """Hang cho xuat kho + quan ly kho (canh bao HSD/ton thap) + nhap lo."""
    today = timezone.now().date()
    if request.method == 'POST':
        form = LoVacXinForm(request.POST)
        if form.is_valid():
            lo = form.save(commit=False)
            lo.nguoi_nhap = request.user
            lo.save()
            return redirect('thu_kho_dashboard')
    else:
        form = LoVacXinForm(initial={'ngay_nhap': today})
    cho_xuat = (QuyTrinhTiem.objects
                .filter(giai_doan=QuyTrinhTiem.B_XUAT_KHO)
                .exclude(lich_hen__trang_thai='huy')
                .select_related('lich_hen__khach_hang', 'vac_xin'))
    # Chi tinh la ton kho khi con tồn > 0 (lo da xuat het khong con la ton kho)
    los = (LoVacXin.objects.filter(so_luong_ton__gt=0)
           .select_related('vac_xin').order_by('han_su_dung'))
    return render(request, 'core/thu_kho/dashboard.html', {
        'cho_xuat': cho_xuat,
        'los': los,
        'sap_het_han': los.filter(han_su_dung__lte=today + timedelta(days=60)),
        'ton_thap': los.filter(so_luong_ton__lt=20),
        'today': today,
        'han_canh_bao': today + timedelta(days=60),
        'form': form,
    })


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_lich_su(request):
    """Lich su xuat kho."""
    ds = (PhieuXuatKho.objects
          .select_related('lo__vac_xin', 'nguoi_xuat',
                          'quy_trinh__lich_hen__khach_hang')
          .order_by('-ngay_xuat'))
    return render(request, 'core/thu_kho/lich_su.html', {'ds': ds})


@trong_nhom('Quan tri', 'Thu kho')
@xframe_options_sameorigin
def phieu_xuat_kho_pdf(request, pk):
    """Phieu xuat kho (PDF) cho 1 lan xuat: bac si xac nhan + thu kho ky."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle

    px = get_object_or_404(PhieuXuatKho.objects.select_related(
        'lo__vac_xin', 'nguoi_xuat', 'nguoi_nhan', 'quy_trinh__lich_hen__khach_hang',
        'quy_trinh__lich_hen__sang_loc__bac_si'), pk=pk)
    font = _font_pdf_dam()

    def vnd(x):
        return format(int(x or 0), ',d')

    def ten(u):
        return (u.get_full_name() or u.username) if u else ''

    lo, vx = px.lo, px.lo.vac_xin
    lh = px.quy_trinh.lich_hen if px.quy_trinh else None
    khach = lh.khach_hang.ho_ten if lh else ''
    sl = getattr(lh, 'sang_loc', None) if lh else None
    bac_si = ten(sl.bac_si) if sl else ''
    thu_kho = ten(px.nguoi_xuat)
    dieu_duong = ten(px.nguoi_nhan)
    now = timezone.localtime(px.ngay_xuat)

    s_ten = ParagraphStyle('sten', fontName=font, fontSize=11)
    s_phai = ParagraphStyle('sphai', fontName=font, fontSize=8.5, alignment=2)
    s_td = ParagraphStyle('std', fontName=font, fontSize=15, alignment=1, spaceBefore=4, spaceAfter=2)
    s_giua = ParagraphStyle('sgiua', fontName=font, fontSize=9, alignment=1, spaceAfter=8)
    s_dong = ParagraphStyle('sdong', fontName=font, fontSize=10, spaceAfter=4)
    s_cell = ParagraphStyle('scell', fontName=font, fontSize=9)
    s_cellc = ParagraphStyle('scellc', fontName=font, fontSize=9, alignment=1)
    s_ky = ParagraphStyle('sky', fontName=font, fontSize=10, alignment=1)
    s_ky_it = ParagraphStyle('skyit', fontName=font, fontSize=8, alignment=1, textColor=colors.grey)
    s_tenky = ParagraphStyle('stenky', fontName=font, fontSize=10, alignment=1, spaceBefore=20)

    el = []
    hdr = Table([[Paragraph('PHÒNG KHÁM TIÊM CHỦNG AN TÂM', s_ten),
                  Paragraph('Mẫu số: 02-VT', s_phai)]], colWidths=[11 * cm, 7 * cm])
    hdr.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    el.append(hdr)
    el.append(Paragraph('PHIẾU XUẤT KHO', s_td))
    el.append(Paragraph(f'Ngày {now:%d/%m/%Y %H:%M}    -    Số: XK-{px.id:05d}', s_giua))

    el.append(Paragraph(f'- Điều dưỡng nhận : {dieu_duong}', s_dong))
    el.append(Paragraph(f'- Lý do xuất kho : Tiêm cho khách {khach} (theo chỉ định bác sĩ)', s_dong))
    el.append(Paragraph('- Xuất tại kho : Phòng khám Tiêm chủng An Tâm', s_dong))

    nsx_txt = f'{lo.ngay_san_xuat:%d/%m/%Y}' if lo.ngay_san_xuat else '—'
    head = ['TT', 'Tên vắc-xin', 'Số lô', 'Ngày SX', 'HSD', 'ĐVT', 'Số lượng']
    data = [[Paragraph(h, s_cellc) for h in head],
            ['1', Paragraph(vx.ten, s_cell), Paragraph(lo.so_lo, s_cell),
             nsx_txt, f'{lo.han_su_dung:%d/%m/%Y}', 'Liều', str(px.so_luong)],
            ['', Paragraph('Tổng cộng', s_cell), '', '', '', '', str(px.so_luong)]]
    t = Table(data, colWidths=[1.2 * cm, 4.8 * cm, 2.6 * cm, 2.6 * cm, 2.6 * cm, 1.6 * cm, 2.6 * cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e9ecef')),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#adb5bd')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    el.append(t)
    el.append(Spacer(1, 0.5 * cm))

    cot_bs = [Paragraph('Bác sĩ xác nhận', s_ky), Paragraph('(ký, ghi rõ họ tên)', s_ky_it),
              Paragraph(bac_si, s_tenky)]
    cot_nhan = [Paragraph('Điều dưỡng', s_ky), Paragraph('(ký, ghi rõ họ tên)', s_ky_it),
                Paragraph(dieu_duong, s_tenky)]
    cot_tk = [Paragraph('Thủ kho', s_ky), Paragraph('(ký, ghi rõ họ tên)', s_ky_it),
              Paragraph(thu_kho, s_tenky)]
    t_ky = Table([[cot_bs, cot_nhan, cot_tk]], colWidths=[6 * cm, 6 * cm, 6 * cm])
    t_ky.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    el.append(t_ky)

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=A4, topMargin=1.3 * cm, bottomMargin=1.3 * cm,
                      leftMargin=1.5 * cm, rightMargin=1.5 * cm).build(el)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    cach = 'attachment' if request.GET.get('tai') else 'inline'
    resp['Content-Disposition'] = f'{cach}; filename=phieu_xuat_kho_{px.id}.pdf'
    return resp


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_nhap_lich_su(request):
    """Lich su nhap kho: moi lo = 1 lan nhap. SL nhap = ton + da xuat + da huy."""
    los = (LoVacXin.objects.select_related('vac_xin', 'nguoi_nhap')
           .order_by('-ngay_nhap', '-id'))
    ds = []
    for lo in los:
        da_xuat = lo.phieu_xuat.aggregate(s=Sum('so_luong'))['s'] or 0
        da_huy = lo.phieu_huy.aggregate(s=Sum('so_luong'))['s'] or 0
        ds.append({'lo': lo, 'sl_nhap': lo.so_luong_ton + da_xuat + da_huy})
    return render(request, 'core/thu_kho/nhap_lich_su.html', {'ds': ds})


def _excel_ngay(v):
    """Chuan hoa o ngay trong Excel ve date / chuoi cho form."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return str(v).strip()


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_excel_mau(request):
    """Tai file Excel mau de nhap kho hang loat."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NhapKho'
    ws.append(['Tên vắc-xin', 'Số lô', 'Ngày sản xuất', 'Ngày nhập', 'Hạn sử dụng', 'Số lượng'])
    vx = VacXin.objects.first()
    ws.append([vx.ten if vx else 'Ví dụ: Vaxigrip Tetra',
               'LO-001', '2026-01-01', '2026-06-01', '2027-06-01', 100])
    for col, w in zip('ABCDEF', (28, 14, 14, 14, 14, 10)):
        ws.column_dimensions[col].width = w
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=mau_nhap_kho.xlsx'
    wb.save(resp)
    return resp


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_nhap_excel(request):
    """Nhap kho hang loat tu file Excel (.xlsx)."""
    ket_qua = None
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            wb = openpyxl.load_workbook(request.FILES['file'], data_only=True)
        except Exception:
            messages.error(request, 'Không đọc được file. Vui lòng dùng đúng định dạng .xlsx.')
            return redirect('thu_kho_nhap_excel')
        ws = wb.active
        thanh_cong, loi = 0, []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            ten, so_lo, ngay_sx, ngay_nhap, han_sd, so_luong = (list(row) + [None] * 6)[:6]
            if ten is None or not str(ten).strip():
                continue
            vx = VacXin.objects.filter(ten__iexact=str(ten).strip()).first()
            if not vx:
                loi.append(f'Dòng {i}: không tìm thấy vắc-xin "{str(ten).strip()}".')
                continue
            form = LoVacXinForm({
                'vac_xin': vx.id,
                'so_lo': str(so_lo).strip() if so_lo is not None else '',
                'ngay_san_xuat': _excel_ngay(ngay_sx),
                'ngay_nhap': _excel_ngay(ngay_nhap),
                'han_su_dung': _excel_ngay(han_sd),
                'so_luong_ton': so_luong,
            })
            if form.is_valid():
                lo = form.save(commit=False)
                lo.nguoi_nhap = request.user
                lo.save()
                thanh_cong += 1
            else:
                chi_tiet = '; '.join(f'{e[0]}' for e in form.errors.values())
                loi.append(f'Dòng {i}: {chi_tiet}')
        if thanh_cong:
            _ghi_lich_su(request, f'Nhập kho {thanh_cong} lô từ Excel')
            messages.success(request, f'Đã nhập thành công {thanh_cong} lô.')
        if loi:
            messages.warning(request, f'{len(loi)} dòng bị bỏ qua do lỗi.')
        ket_qua = {'thanh_cong': thanh_cong, 'loi': loi}
    return render(request, 'core/thu_kho/nhap_excel.html', {'ket_qua': ket_qua})


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_huy(request):
    """Huy vac-xin (het han/loi chat luong) -> tru ton + ghi phieu huy."""
    loi = None
    if request.method == 'POST':
        lo_id = request.POST.get('lo')
        try:
            so_luong = int(request.POST.get('so_luong', 0))
        except (TypeError, ValueError):
            so_luong = 0
        lo = LoVacXin.objects.filter(pk=lo_id).first()
        if not lo:
            loi = 'Vui lòng chọn lô cần hủy.'
        elif so_luong <= 0:
            loi = 'Số lượng hủy phải lớn hơn 0.'
        elif so_luong > lo.so_luong_ton:
            loi = f'Lô chỉ còn {lo.so_luong_ton} liều.'
        else:
            lo.so_luong_ton -= so_luong
            lo.save()
            PhieuHuyKho.objects.create(
                lo=lo, so_luong=so_luong,
                ly_do=request.POST.get('ly_do', 'het_han'),
                ghi_chu=request.POST.get('ghi_chu', ''),
                nguoi_huy=request.user)
            return redirect('thu_kho_huy')
    today = timezone.now().date()
    los = (LoVacXin.objects.filter(so_luong_ton__gt=0)
           .select_related('vac_xin').order_by('han_su_dung'))
    return render(request, 'core/thu_kho/huy.html', {
        'los': los,
        # Lo can xu ly: da het han hoac sap het han (<= 60 ngay), con ton
        'canh_bao_los': los.filter(han_su_dung__lte=today + timedelta(days=60)),
        'lich_su': PhieuHuyKho.objects.select_related('lo__vac_xin', 'nguoi_huy'),
        'ly_do_choices': PhieuHuyKho.LY_DO,
        'today': today,
        'loi': loi,
    })


def _du_lieu_bao_cao_kho(today):
    """Du lieu bao cao xuat-nhap-ton (dung chung cho HTML va PDF)."""
    rows = []
    for vx in VacXin.objects.all().order_by('ten'):
        ton = vx.lo.aggregate(s=Sum('so_luong_ton'))['s'] or 0
        xuat = (PhieuXuatKho.objects.filter(lo__vac_xin=vx)
                .aggregate(s=Sum('so_luong'))['s'] or 0)
        so_lo = vx.lo.count()
        if so_lo == 0 and xuat == 0:
            continue
        rows.append({'vx': vx, 'so_lo': so_lo, 'ton': ton,
                     'xuat': xuat, 'nhap': ton + xuat})
    return {
        'rows': rows,
        'tong_ton': sum(r['ton'] for r in rows),
        'tong_xuat': sum(r['xuat'] for r in rows),
        'het_han': LoVacXin.objects.filter(
            so_luong_ton__gt=0, han_su_dung__lt=today).select_related('vac_xin'),
        'sap_het': LoVacXin.objects.filter(
            so_luong_ton__gt=0, han_su_dung__gte=today,
            han_su_dung__lte=today + timedelta(days=60)).select_related('vac_xin'),
    }


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_bao_cao(request):
    """Bao cao xuat-nhap-ton theo vac-xin + lo het han."""
    today = timezone.now().date()
    ctx = _du_lieu_bao_cao_kho(today)
    ctx['today'] = today
    return render(request, 'core/thu_kho/bao_cao.html', ctx)


def _font_pdf():
    """Dang ky font Unicode (ho tro tieng Viet) cho PDF, tra ten font."""
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    if 'VN' in pdfmetrics.getRegisteredFontNames():
        return 'VN'
    for p in (r'C:\Windows\Fonts\arial.ttf', r'C:\Windows\Fonts\segoeui.ttf',
              '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'):
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont('VN', p))
                return 'VN'
            except Exception:
                pass
    return 'Helvetica'          # du phong: khong dau nhung khong crash


def _font_pdf_dam():
    """Font Unicode + dang ky bien the dam (de dung <b> trong Paragraph PDF)."""
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    font = _font_pdf()
    if font == 'VN' and 'VN-Bold' not in pdfmetrics.getRegisteredFontNames():
        for bp in (r'C:\Windows\Fonts\arialbd.ttf', r'C:\Windows\Fonts\segoeuib.ttf',
                   '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'):
            if os.path.exists(bp):
                try:
                    pdfmetrics.registerFont(TTFont('VN-Bold', bp))
                    pdfmetrics.registerFontFamily('VN', normal='VN', bold='VN-Bold',
                                                  italic='VN', boldItalic='VN-Bold')
                except Exception:
                    pass
                break
    return font


@trong_nhom('Quan tri', 'Thu kho')
@xframe_options_sameorigin
def thu_kho_bao_cao_pdf(request):
    """Xuat bao cao xuat-nhap-ton ra file PDF."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle

    today = timezone.now().date()
    d = _du_lieu_bao_cao_kho(today)
    font = _font_pdf()

    tieu_de = ParagraphStyle('td', fontName=font, fontSize=15, alignment=1, spaceAfter=4)
    phu = ParagraphStyle('phu', fontName=font, fontSize=9, alignment=1,
                         textColor=colors.grey, spaceAfter=2)
    thuong = ParagraphStyle('th', fontName=font, fontSize=10, spaceAfter=2)

    nguoi_xuat = request.user.get_full_name() or request.user.username
    el = [
        Paragraph('BÁO CÁO XUẤT – NHẬP – TỒN KHO', tieu_de),
        Paragraph('Phòng khám Tiêm chủng An Tâm', phu),
        Paragraph(f'Ngày xuất: {today:%d/%m/%Y}', phu),
        Paragraph(f'Người xuất báo cáo: {nguoi_xuat}', phu),
        Spacer(1, 0.4 * cm),
        Paragraph(f'Tổng tồn kho: {d["tong_ton"]} liều   |   '
                  f'Tổng đã xuất: {d["tong_xuat"]} liều   |   '
                  f'Lô đã hết hạn: {d["het_han"].count()}', thuong),
        Spacer(1, 0.2 * cm),
    ]

    data = [['STT', 'Vắc-xin', 'Số lô', 'Tổng nhập', 'Đã xuất', 'Tồn hiện tại']]
    for i, r in enumerate(d['rows'], 1):
        data.append([str(i), r['vx'].ten, str(r['so_lo']),
                     str(r['nhap']), str(r['xuat']), str(r['ton'])])
    if not d['rows']:
        data.append(['', 'Chưa có dữ liệu kho', '', '', '', ''])

    t = Table(data, repeatRows=1,
              colWidths=[1.2 * cm, 6.5 * cm, 2 * cm, 2.4 * cm, 2.2 * cm, 2.7 * cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#adb5bd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    el.append(t)

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                      leftMargin=1.5 * cm, rightMargin=1.5 * cm).build(el)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    # ?tai=1 -> tai ve; mac dinh inline de xem truoc trong trinh duyet
    cach = 'attachment' if request.GET.get('tai') else 'inline'
    resp['Content-Disposition'] = f'{cach}; filename=bao_cao_kho_{today:%Y%m%d}.pdf'
    return resp


@trong_nhom('Quan tri', 'Thu kho')
def thu_kho_bao_cao_xem(request):
    """Trang xem truoc bao cao PDF (nhung inline) truoc khi tai xuong."""
    return render(request, 'core/thu_kho/bao_cao_pdf_xem.html')


# ===================== KHU VUC QUAN TRI =====================

def _ghi_lich_su(request, hanh_dong):
    """Ghi lai thao tac chinh sua cua admin."""
    LichSuQuanTri.objects.create(nguoi=request.user, hanh_dong=hanh_dong)


def _bao_le_tan(noi_dung, duong_dan=''):
    """Gui thong bao (chuong) cho tat ca le tan dang hoat dong."""
    nhom = Group.objects.filter(name='Le tan').first()
    if nhom:
        for lt in nhom.user_set.filter(is_active=True):
            ThongBao.objects.create(nguoi_nhan=lt, noi_dung=noi_dung,
                                    duong_dan=duong_dan)


def _canh_bao_phac_do_tuoi(request, pd):
    """Canh bao mem khi vac-xin trong phac do lech nhom tuoi (khong chan luu)."""
    canh_bao = []
    for ct in pd.chi_tiet.select_related('vac_xin'):
        vx = ct.vac_xin
        if pd.nhom == 'tre_em' and vx.do_tuoi_min_thang >= TUOI_NGUOI_LON_MIN_THANG:
            canh_bao.append(f'"{vx.ten}" chỉ dành cho người lớn '
                            f'(từ {vx.do_tuoi_min_thang} tháng)')
        elif pd.nhom == 'nguoi_lon' and vx.do_tuoi_max_thang <= TUOI_TRE_EM_MAX_THANG:
            canh_bao.append(f'"{vx.ten}" chỉ dành cho trẻ em '
                            f'(đến {vx.do_tuoi_max_thang} tháng)')
    if canh_bao:
        messages.warning(request,
                         'Lưu ý độ tuổi không khớp nhóm phác đồ: '
                         + '; '.join(canh_bao) + '.')


@trong_nhom('Quan tri')
def quan_tri_tai_khoan(request):
    """Tao tai khoan nhan vien + phan quyen."""
    if request.method == 'POST':
        form = TaiKhoanNhanVienForm(request.POST)
        if form.is_valid():
            nv = form.save()
            _ghi_lich_su(request, f'Tạo tài khoản "{nv.username}" '
                                  f'(vai trò {form.cleaned_data["vai_tro"]})')
            messages.success(request, f'Đã tạo tài khoản "{nv.username}".')
            return redirect('quan_tri_tai_khoan')
    else:
        form = TaiKhoanNhanVienForm()
    vai_tro = request.GET.get('vai_tro', '')         # '' chua chon, 'tat_ca', hoac ten nhom
    trang_thai = request.GET.get('trang_thai', '')   # '' = tat ca, 'hoat_dong', 'da_khoa'
    if vai_tro == 'tat_ca':
        nhan_vien = (User.objects.filter(is_staff=True)
                     .prefetch_related('groups').order_by('username'))
    elif vai_tro:
        nhan_vien = (User.objects.filter(is_staff=True, groups__name=vai_tro)
                     .prefetch_related('groups').order_by('username'))
    else:
        nhan_vien = User.objects.none()      # chua chon vai tro -> bang trong
    if vai_tro:
        if trang_thai == 'hoat_dong':
            nhan_vien = nhan_vien.filter(is_active=True)
        elif trang_thai == 'da_khoa':
            nhan_vien = nhan_vien.filter(is_active=False)
    return render(request, 'core/quan_tri/tai_khoan.html', {
        'form': form, 'nhan_vien': nhan_vien,
        'vai_tro': vai_tro, 'vai_tro_list': TaiKhoanNhanVienForm.VAI_TRO,
        'trang_thai': trang_thai})


VAI_TRO_NHOM = [c[0] for c in TaiKhoanNhanVienForm.VAI_TRO]


@trong_nhom('Quan tri')
def quan_tri_tai_khoan_sua(request, pk):
    """Sua thong tin + vai tro + mat khau nhan vien."""
    nv = get_object_or_404(User, pk=pk, is_staff=True)
    if nv.is_superuser:
        messages.warning(request, 'Không thể sửa tài khoản quản trị cấp cao.')
        return redirect('quan_tri_tai_khoan')
    if request.method == 'POST':
        form = SuaNhanVienForm(request.POST)
        if form.is_valid():
            nv.first_name = form.cleaned_data['ho_ten']
            mk = form.cleaned_data.get('mat_khau_moi')
            if mk:
                nv.set_password(mk)
            nv.save()
            # Cap nhat vai tro: bo cac nhom vai tro cu, them vai tro moi
            nv.groups.remove(*nv.groups.filter(name__in=VAI_TRO_NHOM))
            nhom, _ = Group.objects.get_or_create(name=form.cleaned_data['vai_tro'])
            nv.groups.add(nhom)
            _ghi_lich_su(request, f'Sửa tài khoản "{nv.username}" '
                                  f'(vai trò {form.cleaned_data["vai_tro"]})')
            messages.success(request, f'Đã cập nhật tài khoản "{nv.username}".')
            return redirect('quan_tri_tai_khoan')
    else:
        vai_tro_hien = nv.groups.values_list('name', flat=True).first() or 'Le tan'
        form = SuaNhanVienForm(initial={'ho_ten': nv.first_name, 'vai_tro': vai_tro_hien})
    return render(request, 'core/quan_tri/tai_khoan_sua.html', {'form': form, 'nv': nv})


@trong_nhom('Quan tri')
def quan_tri_tai_khoan_khoa(request, pk):
    """Khoa / mo khoa tai khoan nhan vien (is_active)."""
    nv = get_object_or_404(User, pk=pk, is_staff=True)
    if request.method == 'POST' and not nv.is_superuser and nv != request.user:
        nv.is_active = not nv.is_active
        nv.save()
        trang_thai = 'Khóa' if not nv.is_active else 'Mở khóa'
        _ghi_lich_su(request, f'{trang_thai} tài khoản "{nv.username}"')
        messages.success(request, f'Đã {trang_thai.lower()} "{nv.username}".')
    return redirect('quan_tri_tai_khoan')


@trong_nhom('Quan tri')
def quan_tri_vac_xin(request):
    """Them vac-xin moi (de nhap hang)."""
    if request.method == 'POST':
        form = VacXinForm(request.POST)
        if form.is_valid():
            vx = form.save()
            _ghi_lich_su(request, f'Thêm vắc-xin "{vx.ten}"')
            messages.success(request, f'Đã thêm vắc-xin "{vx.ten}".')
            return redirect('quan_tri_vac_xin')
    else:
        form = VacXinForm()
    ds = VacXin.objects.all().order_by('ten')
    return render(request, 'core/quan_tri/vac_xin.html', {'form': form, 'ds': ds})


@trong_nhom('Quan tri')
def quan_tri_vac_xin_sua(request, pk):
    """Sua vac-xin (mo khi bam dup vao dong)."""
    vx = get_object_or_404(VacXin, pk=pk)
    if request.method == 'POST':
        form = VacXinForm(request.POST, instance=vx)
        if form.is_valid():
            form.save()
            _ghi_lich_su(request, f'Sửa vắc-xin "{vx.ten}"')
            messages.success(request, f'Đã cập nhật "{vx.ten}".')
            return redirect('quan_tri_vac_xin')
    else:
        form = VacXinForm(instance=vx)
    return render(request, 'core/quan_tri/vac_xin_sua.html', {'form': form, 'vx': vx})


@trong_nhom('Quan tri')
def quan_tri_vac_xin_xoa(request, pk):
    """Xoa vac-xin (chan neu da co mui tiem su dung)."""
    vx = get_object_or_404(VacXin, pk=pk)
    if request.method == 'POST':
        ten = vx.ten
        try:
            vx.delete()
            _ghi_lich_su(request, f'Xóa vắc-xin "{ten}"')
            messages.success(request, f'Đã xóa vắc-xin "{ten}".')
        except ProtectedError:
            messages.warning(request,
                             f'Không thể xóa "{ten}" vì đang có mũi tiêm sử dụng.')
    return redirect('quan_tri_vac_xin')


@trong_nhom('Quan tri')
def quan_tri_phac_do(request):
    """Danh muc phac do + tao moi (phan master)."""
    if request.method == 'POST':
        form = PhacDoForm(request.POST)
        if form.is_valid():
            pd = form.save()
            _ghi_lich_su(request, f'Thêm phác đồ "{pd.ten}"')
            messages.success(request,
                             f'Đã tạo phác đồ "{pd.ten}". Thêm các mũi tiêm bên dưới.')
            return redirect('quan_tri_phac_do_sua', pk=pd.pk)
    else:
        form = PhacDoForm()
    ds = PhacDo.objects.annotate(so_mui=Count('chi_tiet')).order_by('nhom', 'ten')
    return render(request, 'core/quan_tri/phac_do.html', {'form': form, 'ds': ds})


@trong_nhom('Quan tri')
def quan_tri_phac_do_sua(request, pk):
    """Sua phac do: thong tin chung + cac mui (vac-xin + khoang cach)."""
    pd = get_object_or_404(PhacDo, pk=pk)
    if request.method == 'POST':
        form = PhacDoForm(request.POST, instance=pd)
        formset = PhacDoChiTietFormSet(request.POST, instance=pd)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            _canh_bao_phac_do_tuoi(request, pd)
            _ghi_lich_su(request, f'Sửa phác đồ "{pd.ten}"')
            messages.success(request, f'Đã lưu phác đồ "{pd.ten}".')
            return redirect('quan_tri_phac_do')
    else:
        form = PhacDoForm(instance=pd)
        formset = PhacDoChiTietFormSet(instance=pd)
    return render(request, 'core/quan_tri/phac_do_sua.html',
                  {'form': form, 'formset': formset, 'pd': pd})


@trong_nhom('Quan tri')
def quan_tri_phac_do_xoa(request, pk):
    """Xoa phac do (cac mui chi tiet xoa theo cascade)."""
    pd = get_object_or_404(PhacDo, pk=pk)
    if request.method == 'POST':
        ten = pd.ten
        try:
            pd.delete()
            _ghi_lich_su(request, f'Xóa phác đồ "{ten}"')
            messages.success(request, f'Đã xóa phác đồ "{ten}".')
        except ProtectedError:
            messages.warning(request,
                             f'Không thể xóa "{ten}" vì đang được sử dụng.')
    return redirect('quan_tri_phac_do')


@trong_nhom('Quan tri')
def quan_tri_lich_su(request):
    """Lich su chinh sua cua admin."""
    ds = LichSuQuanTri.objects.select_related('nguoi')[:200]
    return render(request, 'core/quan_tri/lich_su.html', {'ds': ds})


@trong_nhom('Quan tri')
def quan_tri_yeu_cau_mat_khau(request):
    """Danh sach yeu cau dat lai mat khau tu nguoi dung."""
    ds = (YeuCauDatLaiMatKhau.objects
          .select_related('user').prefetch_related('user__groups'))
    return render(request, 'core/quan_tri/yeu_cau_mat_khau.html', {'ds': ds})


@trong_nhom('Quan tri')
def quan_tri_yeu_cau_xu_ly(request, pk):
    """Admin dat lai mat khau cho tai khoan trong yeu cau."""
    yc = get_object_or_404(YeuCauDatLaiMatKhau, pk=pk)
    if request.method == 'POST':
        mk = (request.POST.get('mat_khau_moi') or '').strip()
        if len(mk) < 6:
            messages.error(request, 'Mật khẩu mới phải từ 6 ký tự trở lên.')
        else:
            u = yc.user
            u.set_password(mk)
            u.save()
            yc.da_xu_ly = True
            yc.ngay_xu_ly = timezone.now()
            yc.save()
            if u.is_active:
                ThongBao.objects.create(
                    nguoi_nhan=u,
                    noi_dung='Mật khẩu của bạn đã được quản trị đặt lại. '
                             'Vui lòng đăng nhập với mật khẩu mới.',
                    duong_dan='')
            _ghi_lich_su(request, f'Đặt lại mật khẩu cho "{u.username}"')
            _bao_le_tan(f'Mật khẩu tài khoản "{u.username}" đã được quản trị đặt lại.')
            messages.success(request, f'Đã đặt lại mật khẩu cho "{u.username}".')
    return redirect('quan_tri_yeu_cau_mat_khau')


@trong_nhom('Quan tri')
def quan_tri_yeu_cau_tu_choi(request, pk):
    """Admin tu choi yeu cau dat lai mat khau."""
    yc = get_object_or_404(YeuCauDatLaiMatKhau, pk=pk)
    if request.method == 'POST' and not yc.da_xu_ly:
        yc.da_xu_ly = True
        yc.tu_choi = True
        yc.ngay_xu_ly = timezone.now()
        yc.save()
        if yc.user.is_active:
            ThongBao.objects.create(
                nguoi_nhan=yc.user,
                noi_dung='Yêu cầu đặt lại mật khẩu của bạn đã bị từ chối. '
                         'Vui lòng liên hệ phòng khám nếu cần hỗ trợ.',
                duong_dan='')
        _ghi_lich_su(request, f'Từ chối yêu cầu đặt lại mật khẩu của "{yc.user.username}"')
        _bao_le_tan(f'Yêu cầu đặt lại mật khẩu của "{yc.user.username}" đã bị từ chối.')
        messages.info(request, f'Đã từ chối yêu cầu của "{yc.user.username}".')
    return redirect('quan_tri_yeu_cau_mat_khau')
